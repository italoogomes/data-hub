"""
MMarra Data Hub - Smart Agent v3 (Scoring + LLM Classifier)

3 camadas:
- Layer 1: Scoring por palavras-chave (0ms) - resolve 90%+ dos casos
- Layer 2: LLM Classifier (1-3s) - so quando scoring e ambiguo
- Layer 3: Fallback - sugestoes uteis

Entidades (marcas, empresas, compradores) carregadas do banco automaticamente.

Fluxo:
    Pergunta -> Score (0ms) -> [se ambiguo: LLM classifica (1-3s)] -> SQL template -> Sankhya -> Python formata -> Resposta
"""

import re
import os
import json
import time
import asyncio
import httpx
import requests as req_sync
from pathlib import Path
from datetime import datetime
from typing import Optional

from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).parent.parent.parent
load_dotenv(PROJECT_ROOT / ".env")

from src.llm.query_executor import SafeQueryExecutor
from src.llm.knowledge_base import KnowledgeBase, score_knowledge
from src.llm.alias_resolver import AliasResolver
from src.llm.query_logger import QueryLogger, generate_auto_tags
from src.llm.result_validator import ResultValidator, build_result_data_summary
from src.llm.knowledge_compiler import COMPILED_PATH

# Config
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
LLM_CLASSIFIER_MODEL = os.getenv("LLM_CLASSIFIER_MODEL", os.getenv("LLM_MODEL", "qwen3:4b"))
USE_LLM_CLASSIFIER = os.getenv("USE_LLM_CLASSIFIER", "true").lower() in ("true", "1", "yes")
LLM_CLASSIFIER_TIMEOUT = int(os.getenv("LLM_CLASSIFIER_TIMEOUT", "60"))
USE_LLM_NARRATOR = os.getenv("USE_LLM_NARRATOR", "true").lower() in ("true", "1", "yes")

# Groq API
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.1-8b-instant")
GROQ_TIMEOUT = int(os.getenv("GROQ_TIMEOUT", "10"))
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"


# ============================================================
# GROQ KEY POOL - Rotacao round-robin com cooldown automatico
# ============================================================

class GroqKeyPool:
    """Pool de chaves Groq com rotacao round-robin e cooldown automatico."""

    def __init__(self, keys: list, name: str = "default"):
        self._keys = [k.strip() for k in keys if k.strip()]
        self._name = name
        self._index = 0
        self._cooldown = {}   # {key: timestamp_libera}
        self._usage = {k: 0 for k in self._keys}
        self._daily_usage = {k: 0 for k in self._keys}
        self._last_reset = time.time()
        self._errors = {k: 0 for k in self._keys}
        if self._keys:
            print(f"[GROQ:{name}] Pool inicializado com {len(self._keys)} chave(s)")
        else:
            print(f"[GROQ:{name}] Pool VAZIO - sem chaves configuradas")

    @property
    def available(self) -> bool:
        return len(self._keys) > 0

    def get_key(self) -> str | None:
        """Retorna proxima chave disponivel. Round-robin, pula chaves em cooldown."""
        if not self._keys:
            return None
        self._maybe_reset_daily()
        now = time.time()
        for _ in range(len(self._keys)):
            key = self._keys[self._index % len(self._keys)]
            self._index += 1
            if self._cooldown.get(key, 0) > now:
                continue
            self._usage[key] += 1
            self._daily_usage[key] += 1
            return key
        # Todas em cooldown - retornar a que libera mais cedo
        earliest = min(self._cooldown, key=self._cooldown.get)
        print(f"[GROQ:{self._name}] Todas as chaves em cooldown. Usando ...{earliest[-6:]}")
        return earliest

    def mark_rate_limited(self, key: str, retry_after: int = 60):
        """Marca chave como rate-limited com cooldown."""
        self._cooldown[key] = time.time() + retry_after
        self._errors[key] = self._errors.get(key, 0) + 1
        print(f"[GROQ:{self._name}] ...{key[-6:]} rate-limited, cooldown {retry_after}s")

    def mark_error(self, key: str):
        """Marca erro generico (nao rate limit)."""
        self._errors[key] = self._errors.get(key, 0) + 1

    def _maybe_reset_daily(self):
        """Reseta contadores diarios a meia-noite."""
        now = time.time()
        if now - self._last_reset > 86400:
            self._daily_usage = {k: 0 for k in self._keys}
            self._cooldown = {}
            self._last_reset = now
            print(f"[GROQ:{self._name}] Contadores diarios resetados")

    def stats(self) -> dict:
        return {
            "pool": self._name,
            "keys": len(self._keys),
            "usage_total": {f"...{k[-6:]}": v for k, v in self._usage.items()},
            "usage_today": {f"...{k[-6:]}": v for k, v in self._daily_usage.items()},
            "errors": {f"...{k[-6:]}": v for k, v in self._errors.items() if v > 0},
            "in_cooldown": sum(1 for t in self._cooldown.values() if t > time.time()),
        }


def _make_pool(env_var: str, name: str) -> GroqKeyPool:
    """Cria pool a partir de variavel de ambiente (virgula-separada)."""
    raw = os.getenv(env_var, "")
    keys = [k.strip() for k in raw.split(",") if k.strip()]
    return GroqKeyPool(keys, name)


# Pools globais
pool_classify = _make_pool("GROQ_POOL_CLASSIFY", "classify")
pool_narrate  = _make_pool("GROQ_POOL_NARRATE", "narrate")
pool_train    = _make_pool("GROQ_POOL_TRAIN", "train")

# Fallback: se pools novos nao existem, usar chave unica antiga
if not pool_classify.available:
    _legacy = os.getenv("GROQ_API_KEY", "")
    if _legacy:
        pool_classify = GroqKeyPool([_legacy], "classify-legacy")
        pool_narrate  = GroqKeyPool([_legacy], "narrate-legacy")
        pool_train    = GroqKeyPool([_legacy], "train-legacy")
        print("[GROQ] Usando chave legado GROQ_API_KEY para todos os pools")


async def _groq_request(pool: GroqKeyPool, messages: list, temperature: float = 0.0,
                         max_tokens: int = 400, timeout: int = None) -> dict | None:
    """Faz request ao Groq usando chave do pool, com retry automatico."""
    key = pool.get_key()
    if not key:
        return None

    _timeout = timeout or GROQ_TIMEOUT

    try:
        async with httpx.AsyncClient(timeout=_timeout) as client:
            r = await client.post(
                GROQ_API_URL,
                headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                json={
                    "model": GROQ_MODEL,
                    "messages": messages,
                    "temperature": temperature,
                    "max_tokens": max_tokens,
                }
            )

            if r.status_code == 429:
                retry_after = int(r.headers.get("retry-after", "60"))
                pool.mark_rate_limited(key, retry_after)
                # Tentar com outra chave
                fallback_key = pool.get_key()
                if fallback_key and fallback_key != key:
                    print(f"[GROQ:{pool._name}] Retry com chave alternativa")
                    r = await client.post(
                        GROQ_API_URL,
                        headers={"Authorization": f"Bearer {fallback_key}", "Content-Type": "application/json"},
                        json={"model": GROQ_MODEL, "messages": messages,
                              "temperature": temperature, "max_tokens": max_tokens}
                    )
                    if r.status_code == 429:
                        pool.mark_rate_limited(fallback_key, int(r.headers.get("retry-after", "60")))
                        return None
                else:
                    return None

            r.raise_for_status()
            data = r.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            return {"content": content, "usage": data.get("usage", {})}

    except httpx.TimeoutException:
        print(f"[GROQ:{pool._name}] Timeout ({_timeout}s)")
        pool.mark_error(key)
        return None
    except Exception as e:
        print(f"[GROQ:{pool._name}] Erro: {e}")
        pool.mark_error(key)
        return None


# ============================================================
# SCORING SYSTEM - Palavras-chave e seus pesos por intent
# ============================================================

# Cada palavra/stem ganha pontos pro intent correspondente
# Score >= THRESHOLD = intent detectado

INTENT_SCORES = {
    "pendencia_compras": {
        # Palavras fortes (qualquer uma quase garante o intent)
        "pendencia": 10, "pendencias": 10, "pendente": 10, "pendentes": 10,
        "pend": 8,
        # Contexto compras
        "compra": 6, "compras": 6, "pedido": 5, "pedidos": 5,
        "aberto": 5, "abertos": 5, "atraso": 5, "atrasado": 5, "atrasados": 5,
        # Entidades (reforco)
        "marca": 3, "fornecedor": 3, "comprador": 3, "empresa": 3, "filial": 3,
        # Interrogativos
        "quantos": 2, "quais": 2, "qual": 2, "quanto": 2,
        "tem": 1, "temos": 1, "total": 2, "valor": 2,
        # Verbos
        "falta": 4, "faltam": 4, "faltando": 4, "chegar": 3, "chegando": 3,
        "entrega": 3, "entregar": 3, "previsao": 3,
        # Tipos de compra
        "casada": 6, "casadas": 6, "empenho": 6, "empenhado": 6, "empenhados": 6,
        "vinculada": 5, "vinculado": 5,
        "reposicao": 5, "futura": 4,
        # Perguntas sobre pessoas
        "quem": 4, "responsavel": 4,
        "fornece": 5,
    },
    "estoque": {
        "estoque": 10, "saldo": 8, "disponivel": 6,
        "produto": 3, "peca": 3, "pecas": 3, "item": 2,
        "critico": 5, "baixo": 4, "zerado": 5, "minimo": 4,
        "acabando": 5, "faltando": 4,
        "quantos": 2, "quanto": 2, "quais": 2,
        "codigo": 2, "cod": 2,
        # Referencia/fabricante (reforco - roteamento real e via detect_product_query)
        "referencia": 4, "fabricante": 4, "similar": 5, "similares": 5,
        "equivalente": 5, "equivalentes": 5, "crossref": 5,
        # Aplicacao/veiculo (roteamento real e via detect_product_query)
        "aplicacao": 5, "aplica": 5, "serve": 4, "compativel": 4,
        "veiculo": 3, "motor": 3, "caminhao": 3,
        "scania": 3, "mercedes": 3, "volvo": 3, "vw": 3, "man": 3, "daf": 3, "iveco": 3,
    },
    "vendas": {
        "vendas": 10, "venda": 10, "faturamento": 10,
        "faturou": 8, "vendeu": 8, "vendemos": 8, "faturamos": 8,
        "faturada": 6, "faturadas": 6, "faturado": 6,
        "hoje": 3, "ontem": 3, "semana": 3, "mes": 3, "ano": 3,
        "ticket": 5, "tiquete": 5, "medio": 3,
        "quanto": 3, "total": 3, "valor": 2,
        "ranking": 4, "top": 4, "maiores": 3, "melhores": 3,
        "vendedor": 3, "vendedores": 3,
    },
    "gerar_excel": {
        "excel": 10, "planilha": 10, "xlsx": 10, "csv": 8,
        "arquivo": 8, "download": 8, "baixar": 8, "exportar": 8,
        "gera": 6, "gerar": 6, "gere": 6,
        "relatorio": 5,
    },
    "saudacao": {
        "oi": 10, "ola": 10, "bom": 5, "dia": 3, "boa": 5,
        "tarde": 3, "noite": 3, "fala": 6, "hey": 8, "hello": 8,
        "eai": 8, "eae": 8,
    },
    "ajuda": {
        "ajuda": 10, "help": 10, "menu": 8, "comandos": 8, "opcoes": 8,
        "funciona": 5, "consegue": 5, "funcoes": 6,
    },
}

# Thresholds por intent
INTENT_THRESHOLDS = {
    "pendencia_compras": 8,
    "estoque": 8,
    "vendas": 8,
    "gerar_excel": 8,
    "saudacao": 8,
    "ajuda": 8,
}

# Palavras de confirmacao (para follow-up/excel)
CONFIRM_WORDS = {"sim", "quero", "pode", "claro", "isso", "ok", "beleza", "bora", "vamos", "manda", "faz", "gera", "gere"}

# Score pra decidir se mostra ITENS ou PEDIDOS
VIEW_ITEM_WORDS = {"itens", "item", "produtos", "produto", "detalhe", "detalhes", "detalhado", "lista", "listagem", "peca", "pecas"}
VIEW_ORDER_WORDS = {"pedidos", "pedido", "resumo", "agrupado", "consolidado"}


def normalize(text: str) -> str:
    """Remove acentos e normaliza texto."""
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u',
        'ç': 'c',
    }
    t = text.lower().strip()
    for old, new in replacements.items():
        t = t.replace(old, new)
    return t


def tokenize(text: str) -> list:
    """Extrai palavras normalizadas."""
    return re.findall(r'[a-z0-9]+', normalize(text))


def score_intent(tokens: list) -> dict:
    """Calcula score de cada intent baseado nos tokens (manual + compilado)."""
    scores = {}
    # Manual (prioridade)
    for intent_id, keywords in INTENT_SCORES.items():
        s = 0
        for token in tokens:
            if token in keywords:
                s += keywords[token]
        scores[intent_id] = s
    # Compilado (complementar)
    for intent_id, keywords in _COMPILED_SCORES.items():
        if intent_id not in scores:
            scores[intent_id] = 0
        for token in tokens:
            if token in keywords:
                # So soma se nao foi contado no manual
                if intent_id not in INTENT_SCORES or token not in INTENT_SCORES[intent_id]:
                    scores[intent_id] += keywords[token]
    return scores


def detect_view_mode(tokens: list) -> str:
    """Decide se mostra itens ou pedidos."""
    item_score = sum(1 for t in tokens if t in VIEW_ITEM_WORDS)
    order_score = sum(1 for t in tokens if t in VIEW_ORDER_WORDS)
    if item_score > order_score:
        return "itens"
    return "pedidos"


# ============================================================
# LLM CLASSIFIER (Layer 2 - so quando scoring e ambiguo)
# ============================================================

LLM_CLASSIFIER_PROMPT = """Voce e o interpretador de perguntas do sistema ERP da MMarra Distribuidora Automotiva.
Analise a pergunta do usuario e retorne APENAS um JSON (sem markdown, sem explicacao, sem texto antes ou depois).

# INTENTS POSSIVEIS
- pendencia_compras: pedidos de compra pendentes, o que falta chegar, entregas, previsoes
- estoque: quantidade em estoque, saldo, estoque critico, disponibilidade
- vendas: vendas, faturamento, notas fiscais de venda, receita
- produto: busca por produto especifico, codigo fabricante (HU711/51, WK950/21), similares, cross-reference, visao 360
- conhecimento: como funcionam processos, regras, politicas, explicacoes do ERP
- saudacao: oi, bom dia, ola
- ajuda: o que voce faz, como funciona, help
- desconhecido: nao se encaixa em nenhum

# CAMPOS DO JSON
- intent: um dos intents acima
- marca: nome da marca mencionada (MAIUSCULO) ou null
- fornecedor: nome do fornecedor ou null
- empresa: nome da empresa/filial ou null
- comprador: nome do comprador ou null
- periodo: "hoje"|"ontem"|"semana"|"mes"|"ano" ou null
- view: "pedidos" (agrupado por pedido - padrao para "qual pedido") ou "itens" (produtos individuais - para "qual item/produto")
- filtro: objeto com instrucoes de filtragem ou null (ver abaixo)
- ordenar: campo para ordenar + _DESC ou _ASC (ver campos abaixo)
- top: numero de resultados desejados ou null (ex: 1 para "qual o...", 5 para "top 5")
- tipo_compra: "casada"|"estoque" ou null (tipo de compra mencionado - casada=empenho/vinculada, estoque=reposicao/entrega futura)
- aplicacao: veiculo/motor/maquina mencionado para busca por aplicacao ou null (ex: "SCANIA R450", "MERCEDES ACTROS", "MOTOR DC13")
- extra_columns: lista de colunas extras que o usuario quer ver no relatorio, ou null
  Colunas possiveis: "EMPRESA", "TIPO_COMPRA", "COMPRADOR", "PREVISAO_ENTREGA", "CONFIRMADO",
  "FORNECEDOR", "UNIDADE", "QTD_PEDIDA", "QTD_ATENDIDA", "VLR_UNITARIO", "DIAS_ABERTO",
  "NUM_FABRICANTE", "NUM_ORIGINAL", "REFERENCIA", "APLICACAO"
  Detecte quando o usuario pede para ADICIONAR/VER campos com frases como:
  "contendo X", "com o campo X", "incluindo X", "mostrando X",
  "precisa ter X", "quero ver X tambem", "adiciona X"
  Mapeamentos importantes:
  "codigo fabricante"/"numero fabricante"/"ref fabricante"/"fabricante" = "NUM_FABRICANTE"
  "numero original"/"original" = "NUM_ORIGINAL"
  "referencia"/"ref interna" = "REFERENCIA"
  "previsao"/"previsao de entrega"/"quando chega" = "PREVISAO_ENTREGA"
  "tipo de compra"/"casada ou estoque" = "TIPO_COMPRA"
  "dias"/"dias aberto"/"dias pendente" = "DIAS_ABERTO"
  "comprador"/"quem compra" = "COMPRADOR"
  "empresa"/"filial" = "EMPRESA"
  Se o usuario NAO pediu colunas extras, retorne null.

# FILTRO - como interpretar pedidos de filtragem
O campo "filtro" permite filtrar os dados retornados. Formato: {"campo": "NOME_DO_CAMPO", "operador": "tipo", "valor": "X"}

Operadores:
- "igual": campo == valor (ex: STATUS_ENTREGA == "ATRASADO")
- "vazio": campo esta vazio/nulo (ex: PREVISAO_ENTREGA sem data)
- "nao_vazio": campo tem valor preenchido
- "maior": campo > valor (numerico)
- "menor": campo < valor (numerico)
- "contem": campo contem texto

# CAMPOS DISPONIVEIS (pendencia_compras) - ATENCAO AOS NOMES:
- PEDIDO: numero do pedido de compra
- DT_PEDIDO: data em que o pedido foi feito (quando compramos)
- PREVISAO_ENTREGA: data prevista para o fornecedor entregar (quando vai chegar)
- CONFIRMADO: se o fornecedor confirmou (S/N)
- STATUS_ENTREGA: situacao da entrega (ATRASADO/NO PRAZO/PROXIMO/SEM PREVISAO)
- DIAS_ABERTO: quantos dias o pedido esta em aberto sem receber
- VLR_PENDENTE: valor em reais do que falta receber
- QTD_PENDENTE: quantidade de pecas que falta receber
- FORNECEDOR, CODPROD, PRODUTO, MARCA, QTD_PEDIDA, VLR_UNITARIO, EMPRESA, COMPRADOR
- TIPO_COMPRA: tipo do pedido de compra ("Casada" = empenho/vinculada a venda, "Estoque" = reposicao geral)
- NUM_FABRICANTE: codigo que o fabricante da a peca (AD_NUMFABRICANTE)
- NUM_ORIGINAL: numero original da peca (AD_NUMORIGINAL)
- REFERENCIA: referencia interna do cadastro (PRO.REFERENCIA)

IMPORTANTE - Diferencie corretamente:
- "data de entrega" / "previsao de entrega" / "quando vai chegar" = PREVISAO_ENTREGA (NAO e DT_PEDIDO!)
- "data do pedido" / "quando foi pedido" / "quando comprou" = DT_PEDIDO
- "mais atrasado" / "mais tempo aberto" = DIAS_ABERTO_DESC
- "mais caro" / "maior valor" = VLR_PENDENTE_DESC

Campos para estoque: CODPROD, PRODUTO, MARCA, ESTOQUE_TOTAL, ESTOQUE_MINIMO, CUSTO_MEDIO
Campos para vendas: NUNOTA, CLIENTE, PRODUTO, QTD, VLR_TOTAL, DT_VENDA

# EXEMPLOS
Pergunta: "o que falta chegar da mann?"
{"intent":"pendencia_compras","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "qual pedido esta sem previsao de entrega da tome?"
{"intent":"pendencia_compras","marca":"TOME","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"PREVISAO_ENTREGA","operador":"vazio","valor":null},"ordenar":null,"top":1,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "pedidos atrasados da Donaldson"
{"intent":"pendencia_compras","marca":"DONALDSON","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"STATUS_ENTREGA","operador":"igual","valor":"ATRASADO"},"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "qual pedido da sabo tem a maior previsao de entrega?"
{"intent":"pendencia_compras","marca":"SABO","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":"PREVISAO_ENTREGA_DESC","top":1,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "qual pedido da mann foi feito mais recentemente?"
{"intent":"pendencia_compras","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":"DT_PEDIDO_DESC","top":1,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "qual o pedido mais caro da Mann?"
{"intent":"pendencia_compras","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":"VLR_PENDENTE_DESC","top":1,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "quais pedidos estao confirmados?"
{"intent":"pendencia_compras","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"CONFIRMADO","operador":"igual","valor":"S"},"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "qual item com maior quantidade pendente da Tome?"
{"intent":"pendencia_compras","marca":"TOME","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"itens","filtro":null,"ordenar":"QTD_PENDENTE_DESC","top":1,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "tem algum pedido acima de 50 mil reais?"
{"intent":"pendencia_compras","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"VLR_PENDENTE","operador":"maior","valor":"50000"},"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "vendas de hoje"
{"intent":"vendas","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":"hoje","view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "como funciona a compra casada?"
{"intent":"conhecimento","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "quais pedidos casados da sabo?"
{"intent":"pendencia_compras","marca":"SABO","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"TIPO_COMPRA","operador":"igual","valor":"Casada"},"ordenar":null,"top":null,"tipo_compra":"casada","aplicacao":null,"extra_columns":null}

Pergunta: "quem compra a marca mann?"
{"intent":"pendencia_compras","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "quem fornece a marca sabo?"
{"intent":"pendencia_compras","marca":"SABO","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "pedidos de empenho atrasados"
{"intent":"pendencia_compras","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"STATUS_ENTREGA","operador":"igual","valor":"ATRASADO"},"ordenar":null,"top":null,"tipo_compra":"casada","aplicacao":null,"extra_columns":null}

Pergunta: "compras de estoque da eaton"
{"intent":"pendencia_compras","marca":"EATON","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":{"campo":"TIPO_COMPRA","operador":"igual","valor":"Estoque"},"ordenar":null,"top":null,"tipo_compra":"estoque","aplicacao":null,"extra_columns":null}

Pergunta: "pendencias da Nakata por Ribeirao Preto"
{"intent":"pendencia_compras","marca":"NAKATA","fornecedor":null,"empresa":"RIBEIR","comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "estoque em Uberlandia"
{"intent":"estoque","marca":null,"fornecedor":null,"empresa":"UBERL","comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "vendas de hoje em Itumbiara"
{"intent":"vendas","marca":null,"fornecedor":null,"empresa":"ITUMBI","comprador":null,"periodo":"hoje","view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "HU711/51"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "tudo sobre o produto 133346"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "similares do produto 133346"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "quem tem o filtro WK 950/21"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":null}

Pergunta: "pecas para scania r450"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":"SCANIA R450","extra_columns":null}

Pergunta: "qual filtro serve no mercedes actros"
{"intent":"produto","marca":null,"fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":"MERCEDES ACTROS","extra_columns":null}

Pergunta: "filtros mann para motor dc13"
{"intent":"produto","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":null,"filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":"MOTOR DC13","extra_columns":null}

Pergunta: "pendencias da nakata por ribeirao contendo codigo do fabricante"
{"intent":"pendencia_compras","marca":"NAKATA","fornecedor":null,"empresa":"RIBEIR","comprador":null,"periodo":null,"view":"itens","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":["NUM_FABRICANTE"]}

Pergunta: "itens pendentes da mann mostrando empresa e previsao"
{"intent":"pendencia_compras","marca":"MANN","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"itens","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":["EMPRESA","PREVISAO_ENTREGA"]}

Pergunta: "pendencias da sabo com comprador e tipo de compra"
{"intent":"pendencia_compras","marca":"SABO","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"pedidos","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":["COMPRADOR","TIPO_COMPRA"]}

Pergunta: "me traga as pendencias da nakata incluindo referencia do fabricante e dias em aberto"
{"intent":"pendencia_compras","marca":"NAKATA","fornecedor":null,"empresa":null,"comprador":null,"periodo":null,"view":"itens","filtro":null,"ordenar":null,"top":null,"tipo_compra":null,"aplicacao":null,"extra_columns":["NUM_FABRICANTE","DIAS_ABERTO"]}

Agora classifique:
Pergunta: "{question}"
"""


async def groq_classify(question: str) -> Optional[dict]:
    """Classifica pergunta via Groq usando pool_classify."""
    if not pool_classify.available:
        return None

    prompt = LLM_CLASSIFIER_PROMPT.replace("{question}", question.replace('"', '\\"'))

    result = await _groq_request(
        pool=pool_classify,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0,
        max_tokens=400,
    )

    if not result:
        return None

    content = result["content"].strip()

    # Limpar thinking leak e markdown
    if content.startswith("```"):
        content = content.split("\n", 1)[-1].rsplit("```", 1)[0]
    content = re.sub(r'<think>.*?</think>', '', content, flags=re.DOTALL).strip()

    # Extrair JSON da resposta
    json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)?\}', content)
    if not json_match:
        print(f"[GROQ:classify] JSON nao encontrado: {content[:150]}")
        return None

    try:
        parsed = json.loads(json_match.group())
    except json.JSONDecodeError:
        print(f"[GROQ:classify] JSON invalido: {content[:100]}")
        return None

    print(f"[GROQ:classify] {parsed.get('intent')} | extra_cols={parsed.get('extra_columns')}")

    # Normalizar entidades para MAIUSCULO
    for key in ["marca", "fornecedor", "empresa", "comprador", "aplicacao"]:
        if parsed.get(key):
            parsed[key] = parsed[key].upper().strip()

    return parsed


async def ollama_classify(question: str) -> Optional[dict]:
    """Fallback: Ollama local para classificar. Mais lento mas funciona offline."""
    prompt = LLM_CLASSIFIER_PROMPT.replace("{question}", question.replace('"', '\\"'))

    def _call():
        return req_sync.post(
            f"{OLLAMA_URL}/api/chat",
            json={
                "model": LLM_CLASSIFIER_MODEL,
                "messages": [
                    {"role": "user", "content": prompt},
                    {"role": "assistant", "content": "{"},
                ],
                "stream": False,
                "options": {"temperature": 0, "num_predict": 200, "top_p": 0.1},
            },
            timeout=LLM_CLASSIFIER_TIMEOUT,
        )

    try:
        t0 = time.time()
        resp = await asyncio.to_thread(_call)
        elapsed = time.time() - t0

        if resp.status_code != 200:
            print(f"[OLLAMA-CLS] Erro HTTP {resp.status_code} ({elapsed:.1f}s)")
            return None

        data = resp.json()
        raw = data.get("message", {}).get("content", "").strip()
        raw = "{" + raw  # Prefixo assistant era "{"

        # Limpar thinking leak
        raw = re.sub(r'<think>.*?</think>', '', raw, flags=re.DOTALL).strip()

        # Extrair JSON
        json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)?\}', raw)
        if not json_match:
            print(f"[OLLAMA-CLS] JSON nao encontrado ({elapsed:.1f}s): {raw[:100]}")
            return None

        result = json.loads(json_match.group())
        print(f"[OLLAMA-CLS] OK ({elapsed:.1f}s): {result}")

        for key in ["marca", "fornecedor", "empresa", "comprador"]:
            if result.get(key):
                result[key] = result[key].upper().strip()

        return result

    except req_sync.exceptions.Timeout:
        print(f"[OLLAMA-CLS] Timeout ({LLM_CLASSIFIER_TIMEOUT}s)")
        return None
    except req_sync.exceptions.ConnectionError:
        print(f"[OLLAMA-CLS] Ollama nao acessivel em {OLLAMA_URL}")
        return None
    except json.JSONDecodeError as e:
        print(f"[OLLAMA-CLS] JSON invalido: {e}")
        return None
    except Exception as e:
        print(f"[OLLAMA-CLS] Erro: {type(e).__name__}: {e}")
        return None


async def llm_classify(question: str) -> Optional[dict]:
    """Classificador inteligente: Groq (rapido) -> Ollama (fallback) -> None."""
    if not USE_LLM_CLASSIFIER:
        return None

    # Tentar Groq primeiro (rapido, ~0.5s)
    if pool_classify.available:
        result = await groq_classify(question)
        if result:
            return result
        print("[LLM-CLS] Groq falhou, tentando Ollama...")

    # Fallback: Ollama local
    result = await ollama_classify(question)
    return result


# ============================================================
# LLM NARRATOR - Explica dados de forma natural
# ============================================================

NARRATOR_SYSTEM = """Voce e um assistente de BI da MMarra Distribuidora Automotiva.
Voce recebeu dados de uma consulta ao banco e deve explicar de forma natural e inteligente.
REGRA ABSOLUTA: Responda DIRETAMENTE em portugues brasileiro. NUNCA pense em voz alta. NUNCA comece com Okay, Let me, The user, First, I need. Va direto ao ponto.

REGRAS:
1. Fale como um colega de trabalho experiente - direto, claro, com personalidade
2. Comece respondendo a pergunta principal, depois destaque o que chama atencao
3. Aponte problemas: pedidos atrasados, estoque baixo, valores concentrados
4. Sugira acoes praticas quando fizer sentido (ex: "pode valer ligar pro fornecedor")
5. Use numeros formatados (R$ 864.800, nao 864800.82)
6. Seja conciso - 3 a 6 frases no maximo, nao faca textao
7. Use **negrito** pra destacar numeros e pontos importantes
8. NAO repita os dados em formato de tabela - isso ja foi feito
9. NAO invente dados que nao estao no resumo
10. Se nao tiver nada relevante pra analisar, seja breve"""


async def llm_narrate(question: str, data_summary: str, fallback_response: str) -> str:
    """Pede pro Groq (pool_narrate) explicar os dados de forma natural."""
    if not USE_LLM_NARRATOR or not pool_narrate.available:
        return fallback_response

    user_msg = f"""Pergunta do usuario: "{question}"

Dados retornados do banco:
{data_summary}

Explique esses dados de forma natural e analise o que chama atencao."""

    result = await _groq_request(
        pool=pool_narrate,
        messages=[
            {"role": "system", "content": NARRATOR_SYSTEM},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.6,
        max_tokens=400,
    )

    if not result or not result.get("content"):
        print(f"[NARRATOR] Groq falhou, usando fallback")
        return fallback_response

    text = result["content"].strip()

    # Limpeza minima (Groq nao vaza thinking como Qwen3, mas prevenir)
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL).strip()

    if not text or len(text) < 30:
        return fallback_response

    print(f"[NARRATOR] Groq OK ({len(text)} chars)")
    return text


def build_pendencia_summary(kpis_data: list, detail_data: list, params: dict) -> str:
    """Monta resumo estruturado dos dados de pendencia pra LLM narrar."""
    row = kpis_data[0] if kpis_data and isinstance(kpis_data[0], dict) else {}
    qtd_ped = int(row.get("QTD_PEDIDOS", 0) or 0)
    qtd_itens = int(row.get("QTD_ITENS", 0) or 0)
    vlr = float(row.get("VLR_PENDENTE", 0) or 0)

    lines = [f"Total: {qtd_ped} pedidos, {qtd_itens} itens, R$ {vlr:,.2f}"]

    if params.get("marca"):
        lines.append(f"Filtro: marca {params['marca']}")
    if params.get("fornecedor"):
        lines.append(f"Filtro: fornecedor {params['fornecedor']}")
    if params.get("comprador"):
        lines.append(f"Filtro: comprador {params['comprador']}")

    if detail_data:
        # Status
        status_count = {}
        fornecedores = {}
        total_dias = 0
        max_dias = 0
        for item in detail_data:
            if not isinstance(item, dict):
                continue
            st = item.get("STATUS_ENTREGA", "?")
            status_count[st] = status_count.get(st, 0) + 1

            forn = str(item.get("FORNECEDOR", ""))[:30]
            if forn:
                fornecedores[forn] = fornecedores.get(forn, 0) + float(item.get("VLR_PENDENTE", 0) or 0)

            dias = int(float(item.get("DIAS_ABERTO", 0) or 0))
            total_dias += dias
            max_dias = max(max_dias, dias)

        if status_count:
            status_str = ", ".join(f"{k}: {v}" for k, v in sorted(status_count.items(), key=lambda x: -x[1]))
            lines.append(f"Status: {status_str}")

        if qtd_itens > 0:
            lines.append(f"Media dias aberto: {total_dias // max(qtd_itens, 1)} dias, maximo: {max_dias} dias")

        atrasados = status_count.get("ATRASADO", 0)
        if atrasados > 0:
            pct = round(atrasados / max(qtd_itens, 1) * 100)
            lines.append(f"ATENCAO: {atrasados} itens atrasados ({pct}%)")

        sem_prev = status_count.get("SEM PREVISAO", 0)
        if sem_prev > 0:
            lines.append(f"{sem_prev} itens sem previsao de entrega")

        # Top fornecedores
        if fornecedores:
            top_forn = sorted(fornecedores.items(), key=lambda x: -x[1])[:3]
            forn_str = ", ".join(f"{f}: R$ {v:,.2f}" for f, v in top_forn)
            lines.append(f"Maiores fornecedores: {forn_str}")

    return "\n".join(lines)


def build_vendas_summary(kpi_row: dict, top_vendedores: list, periodo_nome: str) -> str:
    """Monta resumo de vendas pra LLM narrar."""
    qtd = int(kpi_row.get("QTD_VENDAS", 0) or 0)
    fat = float(kpi_row.get("FATURAMENTO", 0) or 0)
    ticket = float(kpi_row.get("TICKET_MEDIO", 0) or 0)

    lines = [
        f"Periodo: {periodo_nome}",
        f"Total: {qtd} notas de venda, faturamento R$ {fat:,.2f}, ticket medio R$ {ticket:,.2f}",
    ]

    if top_vendedores:
        top_str = ", ".join(
            f"{r.get('VENDEDOR','?')}: R$ {float(r.get('FATURAMENTO',0) or 0):,.2f} ({r.get('QTD',0)} notas)"
            for r in top_vendedores[:5] if isinstance(r, dict)
        )
        lines.append(f"Top vendedores: {top_str}")

        # Concentracao
        if len(top_vendedores) >= 2 and fat > 0:
            top1_fat = float(top_vendedores[0].get("FATURAMENTO", 0) or 0)
            pct = round(top1_fat / fat * 100)
            if pct > 40:
                lines.append(f"CONCENTRACAO: {top_vendedores[0].get('VENDEDOR','?')} responde por {pct}% do faturamento")

    return "\n".join(lines)


def build_estoque_summary(data: list, params: dict) -> str:
    """Monta resumo de estoque pra LLM narrar."""
    lines = []
    if params.get("codprod") or params.get("produto_nome"):
        if data:
            row = data[0] if isinstance(data[0], dict) else {}
            est = int(float(row.get("ESTOQUE", 0) or 0))
            est_min = int(float(row.get("ESTMIN", 0) or 0))
            lines.append(f"Produto: {row.get('CODPROD','?')} - {row.get('PRODUTO','?')} ({row.get('MARCA','')})")
            lines.append(f"Estoque: {est} unidades, minimo: {est_min}")
            if est <= est_min:
                lines.append("ALERTA: Estoque abaixo do minimo!")
            if len(data) > 1:
                for r in data:
                    if isinstance(r, dict):
                        lines.append(f"  {r.get('EMPRESA','?')}: {r.get('ESTOQUE',0)} un")
    else:
        lines.append(f"{len(data)} produtos com estoque critico (abaixo do minimo)")
        zerados = sum(1 for r in data if isinstance(r, dict) and int(float(r.get("ESTOQUE", 0) or 0)) == 0)
        if zerados:
            lines.append(f"ALERTA: {zerados} produtos com estoque ZERADO!")
        if data:
            marcas = {}
            for r in data:
                if isinstance(r, dict):
                    m = r.get("MARCA", "SEM MARCA")
                    marcas[m] = marcas.get(m, 0) + 1
            top_marcas = sorted(marcas.items(), key=lambda x: -x[1])[:5]
            lines.append(f"Marcas mais afetadas: {', '.join(f'{m}: {c}' for m, c in top_marcas)}")

    return "\n".join(lines)


# ============================================================
# ENTITY EXTRACTION (scoring-based, not regex-dependent)
# ============================================================

def extract_entities(question: str, known_marcas: set = None, known_empresas: set = None, known_compradores: set = None) -> dict:
    """Extrai entidades da pergunta usando matching com banco."""
    params = {}
    q_upper = question.upper().strip()
    q_norm = normalize(question)
    tokens = tokenize(question)

    # ---- MARCA ----
    # Estrategia 1: "marca X" ou "da X"
    m = re.search(r'(?:MARCA\s+)([A-Z][A-Z0-9\s\.\-&]{1,35}?)(?:\s*[?,!.]|\s+(?:QUE|TEM|TEMOS|EU|TENHO)|\s*$)', q_upper)
    if m:
        params["marca"] = m.group(1).strip()

    if "marca" not in params:
        # DA/DE/DO (standalone) + captura ate verbo/preposicao/pontuacao/fim
        stop_after = r'(?:\s+(?:QUE|TEM|TEMOS|COM|SEM|ESTA|ESTAO|FOI|NAO|PARA|POR|EM|NO|NA|NOS|COMO|ONDE|QUAL|QUAIS|ENTRE|ACIMA|ABAIXO)|\s*[?,!.]|\s*$)'
        m = re.search(r'\b(?:DA|DE|DO|PELA|PELO|PELAS|PELOS)\s+([A-Z][A-Z0-9\s\.\-&]{1,30}?)' + stop_after, q_upper)
        if m:
            candidate = m.group(1).strip()
            noise = {"COMPRA", "COMPRAS", "VENDA", "VENDAS", "EMPRESA", "FORNECEDOR",
                     "MARCA", "PRODUTO", "PRODUTOS", "ESTOQUE", "PEDIDO", "PEDIDOS", "MES",
                     "SEMANA", "ANO", "HOJE", "ONTEM", "PERIODO", "SISTEMA",
                     "TODAS", "TODOS", "TUDO", "GERAL", "MINHA", "MINHAS",
                     "ENTREGA", "PREVISAO", "DATA", "CONFIRMACAO", "COMPRADOR",
                     "VALOR", "QUANTIDADE", "STATUS", "PRAZO", "ATRASO",
                     "ESTA", "ESTAO", "ESSE", "ESSA", "ISSO", "AQUI", "ONDE",
                     "QUEM", "RESPONSAVEL", "FORNECEDORES", "COMPRADORES",
                     "CASADA", "CASADAS", "CASADO", "CASADOS", "EMPENHO",
                     "FUTURA", "REPOSICAO",
                     "FILIAL", "UNIDADE", "LOJA"}
            first_word = candidate.split()[0] if candidate else ""
            if first_word in noise:
                # Buscar ultimo DA/DO (provavelmente antes do nome real)
                last_da = re.search(r'.*\b(?:DA|DO)\s+([A-Z][A-Z0-9\s\.\-&]{1,30}?)' + stop_after, q_upper)
                if last_da:
                    candidate = last_da.group(1).strip()
                    first_word = candidate.split()[0] if candidate else ""
            if candidate not in noise and first_word not in noise and len(candidate) > 1:
                params["marca"] = candidate

    # Estrategia 2: Matching com marcas do banco
    if "marca" not in params and known_marcas:
        # Palavras comuns que NAO devem matchear com marcas
        stop_words = {"DOS", "DAS", "DEL", "UMA", "UNS", "COM", "POR", "QUE",
                       "NAO", "SIM", "MAS", "SEM", "SOB", "TEM", "SAO", "ERA",
                       "FOI", "SER", "TER", "VER", "DAR", "FAZ", "DIZ",
                       "MEU", "SEU", "TEU", "NOS", "VOS", "ELA", "ELE",
                       "PARA", "MAIS", "COMO", "ESSE", "ESSA", "ESTE", "ESTA",
                       "AQUI", "ONDE", "QUAL", "QUEM", "AGORA", "ALEM",
                       "ITENS", "ITEM", "CADA", "DOIS", "TRES", "QUATRO",
                       "PRECISO", "QUERO", "GERAR", "GERA", "TOTAL"}
        # Cada token (e combinacoes de 2) vs banco
        for i, token in enumerate(tokens):
            t_upper = token.upper()
            if len(t_upper) < 3 or t_upper in stop_words:
                continue
            # Match exato
            if t_upper in known_marcas:
                params["marca"] = t_upper
                break
            # Match parcial (token dentro de marca) - minimo 4 chars pra evitar falso positivo
            for m in known_marcas:
                if t_upper in m and len(t_upper) >= 4:
                    params["marca"] = m
                    break
                if m in q_upper and len(m) >= 4:
                    params["marca"] = m
                    break
            if "marca" in params:
                break

    # ---- FORNECEDOR ----
    # "fornecedor Odapel" = filtrar por fornecedor. "fornecedor da SABO" = quem fornece marca (ignorar).
    m = re.search(r'FORNECEDOR\s+([A-Z][A-Z\s\.\-&]{2,40}?)(?:\s*[?,!.]|\s*$)', q_upper)
    if m:
        candidate_forn = m.group(1).strip()
        # Se comeca com DA/DE/DO = provavelmente "fornecedor da [marca]", nao nome de fornecedor
        if not re.match(r'^D[AEOI]\s', candidate_forn):
            params["fornecedor"] = candidate_forn

    # ---- EMPRESA ----
    _CIDADES_EMPRESA = {
        "ARACATUBA": "ARACAT", "ARAÇATUBA": "ARACAT", "ARACAT": "ARACAT",
        "RIBEIRAO PRETO": "RIBEIR", "RIBEIRÃO PRETO": "RIBEIR",
        "RIBEIRAO": "RIBEIR", "RIBEIRÃO": "RIBEIR", "RIBEIR": "RIBEIR",
        "UBERLANDIA": "UBERL", "UBERLÂNDIA": "UBERL", "UBERL": "UBERL",
        "ITUMBIARA": "ITUMBI", "ITUMBI": "ITUMBI",
        "RIO VERDE": "RIO VERDE",
        "GOIANIA": "GOIAN", "GOIÂNIA": "GOIAN", "GOIAN": "GOIAN",
        "SAO JOSE": "SAO JOSE", "SÃO JOSÉ": "SAO JOSE",
    }
    _CIDADES_SET = {c.split()[0] for c in _CIDADES_EMPRESA.keys()}  # {"ARACATUBA","RIBEIRAO",...}
    q_upper_norm = normalize(question).upper()

    def _resolve_cidade(name):
        """Resolve nome/prefixo para prefixo padrao de empresa."""
        for cidade, prefixo in _CIDADES_EMPRESA.items():
            if cidade in name or name in cidade:
                return prefixo
        return name

    # 1. Prefixo explicito: "empresa X", "filial X", "filial de X"
    m = re.search(r'(?:EMPRESA|FILIAL|UNIDADE|LOJA)\s+(?:DE\s+|DA\s+|DO\s+)?([A-Z][A-ZÀ-Ú\s\-]{2,30}?)(?:\s*[?,!.]|\s*$)', q_upper)
    if m:
        params["empresa"] = _resolve_cidade(m.group(1).strip())

    # 2. Preposicao + cidade: "por Ribeirão", "em Uberlândia"
    if "empresa" not in params:
        m_prep = re.search(r'\b(?:POR|EM|PARA|PRA)\s+([A-Z][A-ZÀ-Ú\s]{2,25})', q_upper)
        if m_prep:
            candidate_city = m_prep.group(1).strip()
            resolved = _resolve_cidade(candidate_city)
            if resolved != candidate_city:
                params["empresa"] = resolved

    # 3. Cidade solta (sem preposicao) no texto
    if "empresa" not in params:
        for cidade, prefixo in _CIDADES_EMPRESA.items():
            if cidade in q_upper or cidade in q_upper_norm:
                params["empresa"] = prefixo
                break

    # 4. Match com empresas conhecidas do banco
    if "empresa" not in params and known_empresas:
        # 4a. Com preposicao
        m_prep2 = re.search(r'\b(?:POR|EM|PARA|PRA)\s+([A-Z][A-Z\s\-]{2,30})', q_upper)
        if m_prep2:
            candidate_emp = m_prep2.group(1).strip()
            for emp in known_empresas:
                if candidate_emp in emp or emp in candidate_emp:
                    params["empresa"] = emp
                    break
        # 4b. Match direto
        if "empresa" not in params:
            for emp in known_empresas:
                if emp in q_upper and len(emp) >= 3:
                    params["empresa"] = emp
                    break

    # Limpar marca se pegou cidade por engano (ex: "da Nakata Ribeirao" → marca=NAKATA RIBEIRAO)
    if params.get("marca") and params.get("empresa"):
        marca = params["marca"]
        # Remover sufixo de cidade do nome da marca
        for cidade_word in _CIDADES_SET:
            if marca.endswith(" " + cidade_word) or marca.endswith(" " + cidade_word + " PRETO"):
                params["marca"] = marca[:marca.rfind(" " + cidade_word)].strip()
                break

    # ---- COMPRADOR ----
    # "comprador Ana" = filtrar por comprador. "comprador da SABO" = quem compra marca (ignorar).
    m = re.search(r'COMPRADOR[A]?\s+([A-Z][A-Z\s]{2,25}?)(?:\s*[?,!.]|\s*$)', q_upper)
    if m:
        candidate_comp = m.group(1).strip()
        if not re.match(r'^D[AEOI]\s', candidate_comp):
            params["comprador"] = candidate_comp

    if "comprador" not in params and known_compradores:
        for comp in known_compradores:
            if comp in q_upper and len(comp) >= 3:
                params["comprador"] = comp
                break

    # ---- NUMERO PEDIDO ----
    m = re.search(r'(?:PEDIDO|NOTA|NUNOTA)\s*(?:N(?:UMERO)?\.?)?\s*(\d{4,10})', q_upper)
    if m:
        params["nunota"] = int(m.group(1))

    # ---- CODIGO PRODUTO ----
    m = re.search(r'(?:CODIGO|COD|CODPROD|PRODUTO)\s*(\d{3,8})', q_upper)
    if m:
        params["codprod"] = int(m.group(1))

    # ---- CODIGO FABRICANTE (alfanumerico) ----
    # Detecta codigos de fabricante que NAO sao CODPROD interno
    # Ex: "WK 950/21", "HU711/51", "078115561J", "F026407032"
    if "codprod" not in params:
        # 1. Apos keyword: "referencia WK 950/21", "fabricante 078115561J"
        m = re.search(r'(?:REFERENCIA|FABRICANTE|NUM(?:ERO)?\s*(?:DO\s+)?FAB(?:RICANTE)?|COD(?:IGO)?\s*FAB(?:RICANTE)?|ORIGINAL)\s+([A-Z0-9][A-Z0-9\s\-/\.]{2,30})', q_upper)
        if m:
            params["codigo_fabricante"] = m.group(1).strip()
        # 2. Codigo alfanumerico solto: letras + numeros misturados
        # Ex: "WK 950/21", "HU711/51", "F026407032", "LB13145/3"
        if "codigo_fabricante" not in params:
            m = re.search(r'\b([A-Z]{1,5}\d{2,}[A-Z0-9/\-\.]*)\b', q_upper)
            if m:
                candidate_fab = m.group(1).strip()
                # Evitar falsos positivos com marcas conhecidas curtas
                if len(candidate_fab) >= 4 and not (known_marcas and candidate_fab in known_marcas):
                    params["codigo_fabricante"] = candidate_fab
        # 3. Numero LONGO (7+ digitos) = provavelmente codigo fabricante, nao CODPROD
        if "codigo_fabricante" not in params:
            m = re.search(r'\b(\d{7,15}[A-Z]?)\b', q_upper)
            if m:
                params["codigo_fabricante"] = m.group(1)

    # ---- NOME PRODUTO ----
    m = re.search(r'(?:PRODUTO|PECA|ITEM)\s+([A-Z][A-Z0-9\s\-/]{3,40}?)(?:\s*[?,!.]|\s*$)', q_upper)
    if m and "codprod" not in params and "codigo_fabricante" not in params:
        candidate = m.group(1).strip()
        noise_prod = {"TEM", "TEMOS", "NO", "ESTOQUE", "PENDENTE", "EM", "ABERTO", "SIMILAR", "SIMILARES", "EQUIVALENTE"}
        if candidate not in noise_prod:
            params["produto_nome"] = candidate

    # ---- APLICACAO / VEICULO ----
    if "codprod" not in params and "codigo_fabricante" not in params:
        aplic_match = re.search(
            r'(?:SERVE|APLICA|COMPATIVEL|ENCAIXA|CABE)\s+(?:NO|NA|NOS|NAS|PRA|PARA|COM|EM)\s+([A-Z][A-Z0-9\s\-/]{2,30})',
            q_upper
        )
        if not aplic_match:
            aplic_match = re.search(
                r'(?:PECAS?|PRODUTOS?|FILTROS?)\s+(?:DO|DA|PRO|PRA|PARA|P/)\s+([A-Z][A-Z0-9\s\-/]{2,30})',
                q_upper
            )
        if not aplic_match:
            aplic_match = re.search(
                r'(?:MOTOR|VEICULO|CAMINHAO|ONIBUS|CARRO|MAQUINA)\s+([A-Z][A-Z0-9\s\-/]{2,30})',
                q_upper
            )
        if aplic_match:
            candidate_aplic = aplic_match.group(1).strip()
            noise_aplic = {"ESTOQUE", "PENDENTE", "PENDENCIA", "COMPRA", "COMPRAS",
                           "VENDA", "VENDAS", "MARCA", "EMPRESA", "PRODUTO", "PRODUTOS"}
            first_w = candidate_aplic.split()[0] if candidate_aplic else ""
            if first_w not in noise_aplic and candidate_aplic not in noise_aplic:
                params["aplicacao"] = candidate_aplic

    # ---- PERIODO ----
    q_lower = question.lower()
    if "hoje" in q_lower:
        params["periodo"] = "hoje"
    elif "ontem" in q_lower:
        params["periodo"] = "ontem"
    elif re.search(r'(essa|esta|nessa|nesta)\s*semana', q_lower):
        params["periodo"] = "semana"
    elif re.search(r'(esse|este|nesse|neste)\s*mes', q_norm):
        params["periodo"] = "mes"
    elif re.search(r'mes\s+(passado|anterior)', q_norm):
        params["periodo"] = "mes_passado"
    elif re.search(r'semana\s+(passada|anterior)', q_norm):
        params["periodo"] = "semana_passada"

    return params


# Mapa de prefixo empresa -> nome legivel para exibicao
EMPRESA_DISPLAY = {
    "ARACAT": "Aracatuba",
    "RIBEIR": "Ribeirao Preto",
    "UBERL": "Uberlandia",
    "ITUMBI": "Itumbiara",
    "RIO VERDE": "Rio Verde",
    "GOIAN": "Goiania",
    "SAO JOSE": "Sao Jose do Rio Preto",
}


# ============================================================
# SQL TEMPLATES
# ============================================================

JOINS_PENDENCIA = """
    FROM TGFITE ITE
    JOIN TGFCAB CAB ON CAB.NUNOTA = ITE.NUNOTA
    JOIN TSIEMP EMP ON EMP.CODEMP = ITE.CODEMP
    JOIN TGFPRO PRO ON PRO.CODPROD = ITE.CODPROD
    LEFT JOIN TGFPAR PAR ON PAR.CODPARC = CAB.CODPARC
    LEFT JOIN TGFMAR MAR ON MAR.CODIGO = PRO.CODMARCA
    LEFT JOIN TGFVEN VEN ON VEN.CODVEND = MAR.AD_CODVEND
    LEFT JOIN (
        SELECT V.NUNOTAORIG, V.SEQUENCIAORIG,
               SUM(V.QTDATENDIDA) AS TOTAL_ATENDIDO
        FROM TGFVAR V
        JOIN TGFCAB C ON C.NUNOTA = V.NUNOTA
        WHERE C.STATUSNOTA <> 'C'
        GROUP BY V.NUNOTAORIG, V.SEQUENCIAORIG
    ) V_AGG ON V_AGG.NUNOTAORIG = ITE.NUNOTA
           AND V_AGG.SEQUENCIAORIG = ITE.SEQUENCIA
"""

WHERE_PENDENCIA = """CAB.CODTIPOPER IN (1301, 1313)
      AND CAB.STATUSNOTA <> 'C'
      AND CAB.PENDENTE = 'S'
      AND ITE.PENDENTE = 'S'
      AND (ITE.QTDNEG - NVL(V_AGG.TOTAL_ATENDIDO, 0)) > 0"""


def _safe_sql(value: str) -> str:
    """Sanitiza valor para uso em SQL: remove chars perigosos, escapa aspas."""
    if not value:
        return ""
    s = str(value).strip()
    # Remover chars perigosos de SQL injection
    s = re.sub(r'[;\-\-\/\*\\\x00]', '', s)
    # Escapar aspas simples (Oracle-style)
    s = s.replace("'", "''")
    return s


def _build_where_extra(params, user_context=None):
    w = ""
    if params.get("marca"):
        w += f" AND UPPER(MAR.DESCRICAO) LIKE UPPER('%{_safe_sql(params['marca'])}%')"
    if params.get("fornecedor"):
        w += f" AND UPPER(PAR.NOMEPARC) LIKE UPPER('%{_safe_sql(params['fornecedor'])}%')"
    if params.get("empresa"):
        w += f" AND UPPER(EMP.NOMEFANTASIA) LIKE UPPER('%{_safe_sql(params['empresa'])}%')"
    if params.get("comprador"):
        w += f" AND UPPER(VEN.APELIDO) LIKE UPPER('%{_safe_sql(params['comprador'])}%')"
    if params.get("nunota"):
        w += f" AND CAB.NUNOTA = {int(params['nunota'])}"
    if params.get("codprod"):
        w += f" AND PRO.CODPROD = {int(params['codprod'])}"
    if params.get("produto_nome") and not params.get("codprod"):
        w += f" AND UPPER(PRO.DESCRPROD) LIKE UPPER('%{_safe_sql(params['produto_nome'])}%')"
    if params.get("aplicacao"):
        w += f" AND UPPER(PRO.CARACTERISTICAS) LIKE UPPER('%{_safe_sql(params['aplicacao'])}%')"
    # Tipo de compra (casada/estoque) via CODTIPOPER (mais eficiente que filtrar pos-query)
    tc = (params.get("tipo_compra") or "").upper()
    if tc in ("CASADA", "EMPENHO", "VINCULADA"):
        w += " AND CAB.CODTIPOPER = 1313"
    elif tc in ("ESTOQUE", "FUTURA", "REPOSICAO"):
        w += " AND CAB.CODTIPOPER = 1301"
    if user_context:
        role = user_context.get("role", "vendedor")
        if role in ("admin", "diretor", "ti"):
            pass
        elif role == "gerente":
            team = user_context.get("team_codvends", [])
            if team:
                w += f" AND MAR.AD_CODVEND IN ({','.join(str(c) for c in team)})"
        elif role in ("vendedor", "comprador"):
            codvend = user_context.get("codvend", 0)
            if codvend:
                w += f" AND MAR.AD_CODVEND = {codvend}"
    return w


def sql_pendencia_compras(params, user_context=None):
    we = _build_where_extra(params, user_context)
    sql_kpis = f"""
        SELECT
            COUNT(DISTINCT CAB.NUNOTA) AS QTD_PEDIDOS,
            COUNT(*) AS QTD_ITENS,
            NVL(SUM(ROUND((ITE.QTDNEG - NVL(V_AGG.TOTAL_ATENDIDO, 0)) * ITE.VLRUNIT, 2)), 0) AS VLR_PENDENTE
        {JOINS_PENDENCIA}
        WHERE {WHERE_PENDENCIA} {we}
    """
    sql_detail = f"""
        SELECT
            EMP.NOMEFANTASIA AS EMPRESA,
            CAB.NUNOTA AS PEDIDO,
            CASE WHEN CAB.CODTIPOPER = 1313 THEN 'Casada' WHEN CAB.CODTIPOPER = 1301 THEN 'Estoque' END AS TIPO_COMPRA,
            NVL(VEN.APELIDO, 'SEM COMPRADOR') AS COMPRADOR,
            TO_CHAR(CAB.DTNEG, 'DD/MM/YYYY') AS DT_PEDIDO,
            NVL(TO_CHAR(CAB.DTPREVENT, 'DD/MM/YYYY'), 'Sem previsao') AS PREVISAO_ENTREGA,
            CASE WHEN CAB.STATUSNOTA = 'L' THEN 'Sim' ELSE 'Nao' END AS CONFIRMADO,
            PAR.NOMEPARC AS FORNECEDOR,
            PRO.CODPROD,
            PRO.DESCRPROD AS PRODUTO,
            NVL(MAR.DESCRICAO, '') AS MARCA,
            NVL(PRO.CARACTERISTICAS, '') AS APLICACAO,
            NVL(PRO.AD_NUMFABRICANTE, '') AS NUM_FABRICANTE,
            ITE.CODVOL AS UNIDADE,
            ITE.QTDNEG AS QTD_PEDIDA,
            NVL(V_AGG.TOTAL_ATENDIDO, 0) AS QTD_ATENDIDA,
            (ITE.QTDNEG - NVL(V_AGG.TOTAL_ATENDIDO, 0)) AS QTD_PENDENTE,
            ITE.VLRUNIT AS VLR_UNITARIO,
            ROUND((ITE.QTDNEG - NVL(V_AGG.TOTAL_ATENDIDO, 0)) * ITE.VLRUNIT, 2) AS VLR_PENDENTE,
            TRUNC(SYSDATE) - TRUNC(CAB.DTNEG) AS DIAS_ABERTO,
            CASE
                WHEN CAB.DTPREVENT IS NULL THEN 'SEM PREVISAO'
                WHEN CAB.DTPREVENT < SYSDATE THEN 'ATRASADO'
                WHEN CAB.DTPREVENT < SYSDATE + 7 THEN 'PROXIMO'
                ELSE 'NO PRAZO'
            END AS STATUS_ENTREGA
        {JOINS_PENDENCIA}
        WHERE {WHERE_PENDENCIA} {we}
        ORDER BY
            CASE WHEN CAB.DTPREVENT IS NULL THEN 1 WHEN CAB.DTPREVENT < SYSDATE THEN 0 WHEN CAB.DTPREVENT < SYSDATE + 7 THEN 2 ELSE 3 END,
            MAR.DESCRICAO, PRO.DESCRPROD
    """
    filtro_desc = []
    if params.get("marca"): filtro_desc.append(f"marca **{params['marca']}**")
    if params.get("fornecedor"): filtro_desc.append(f"fornecedor **{params['fornecedor']}**")
    if params.get("empresa"):
        emp_display = EMPRESA_DISPLAY.get(params['empresa'].upper(), params['empresa'])
        filtro_desc.append(f"empresa **{emp_display}**")
    if params.get("comprador"): filtro_desc.append(f"comprador **{params['comprador']}**")
    if params.get("nunota"): filtro_desc.append(f"pedido **{params['nunota']}**")
    if params.get("codprod"): filtro_desc.append(f"produto **{params['codprod']}**")
    if params.get("produto_nome") and not params.get("codprod"): filtro_desc.append(f"produto **{params['produto_nome']}**")
    desc = "pendencia de compras"
    if filtro_desc: desc += " da " + ", ".join(filtro_desc)
    return sql_kpis, sql_detail, desc


def _build_periodo_filter(params, date_col="C.DTNEG"):
    p = params.get("periodo", "mes")
    m = {
        "hoje": f"AND {date_col} >= TRUNC(SYSDATE) AND {date_col} < TRUNC(SYSDATE) + 1",
        "ontem": f"AND {date_col} >= TRUNC(SYSDATE) - 1 AND {date_col} < TRUNC(SYSDATE)",
        "semana": f"AND {date_col} >= TRUNC(SYSDATE, 'IW') AND {date_col} < TRUNC(SYSDATE) + 1",
        "semana_passada": f"AND {date_col} >= TRUNC(SYSDATE, 'IW') - 7 AND {date_col} < TRUNC(SYSDATE, 'IW')",
        "mes": f"AND {date_col} >= TRUNC(SYSDATE, 'MM') AND {date_col} < TRUNC(SYSDATE) + 1",
        "mes_passado": f"AND {date_col} >= ADD_MONTHS(TRUNC(SYSDATE, 'MM'), -1) AND {date_col} < TRUNC(SYSDATE, 'MM')",
        "ano": f"AND {date_col} >= TRUNC(SYSDATE, 'YYYY') AND {date_col} < TRUNC(SYSDATE) + 1",
    }
    return m.get(p, m["mes"])


PERIODO_NOMES = {
    "hoje": "hoje", "ontem": "ontem", "semana": "esta semana",
    "semana_passada": "semana passada", "mes": "este mes",
    "mes_passado": "mes passado", "ano": "este ano",
}


# ============================================================
# FORMATACAO
# ============================================================

def fmt_brl(valor):
    try:
        v = float(valor or 0)
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"

def fmt_num(valor):
    try:
        return f"{int(float(valor or 0)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "0"


# ============================================================
# PRODUTO - Resolucao de codigo fabricante e similares
# ============================================================

SIMILAR_WORDS = {"similar", "similares", "equivalente", "equivalentes",
                 "alternativ", "substitut", "cross", "crossref",
                 "auxiliar", "auxiliares", "outras marcas", "outra marca",
                 "quais marcas", "que marcas"}


def _trunc(text, max_len=40):
    """Trunca texto para caber em tabelas."""
    if not text:
        return ""
    s = str(text).strip()
    return s[:max_len] + "..." if len(s) > max_len else s


def _sanitize_code(code: str) -> str:
    """Remove espacos, tracos e barras para comparacao flexivel de codigos."""
    return re.sub(r'[\s\-/\.]', '', code).upper()


async def resolve_manufacturer_code(code_input: str, executor) -> dict:
    """Busca produto pelo codigo do fabricante em TGFPRO.

    Busca nos campos: REFERENCIA, AD_NUMFABRICANTE, AD_NUMFABRICANTE2, AD_NUMORIGINAL, REFFORN.
    Normaliza: remove espacos, tracos, barras antes de comparar.

    Returns:
        {"found": bool, "products": [...], "code_searched": str}
    """
    safe_code = _safe_sql(code_input)
    clean = _sanitize_code(safe_code)

    sql = f"""SELECT DISTINCT PRO.CODPROD, PRO.DESCRPROD AS PRODUTO,
        NVL(MAR.DESCRICAO, '') AS MARCA,
        PRO.REFERENCIA, PRO.REFFORN,
        PRO.AD_NUMFABRICANTE, PRO.AD_NUMFABRICANTE2, PRO.AD_NUMORIGINAL,
        NVL(PRO.CARACTERISTICAS, '') AS APLICACAO
    FROM TGFPRO PRO
    LEFT JOIN TGFMAR MAR ON MAR.CODIGO = PRO.CODMARCA
    WHERE PRO.ATIVO = 'S'
      AND (
        UPPER(REPLACE(REPLACE(REPLACE(PRO.REFERENCIA,' ',''),'-',''),'/',''))
            LIKE '%{clean}%'
        OR UPPER(REPLACE(REPLACE(PRO.AD_NUMFABRICANTE,' ',''),'-',''))
            LIKE '%{clean}%'
        OR UPPER(REPLACE(REPLACE(PRO.AD_NUMFABRICANTE2,' ',''),'-',''))
            LIKE '%{clean}%'
        OR UPPER(REPLACE(REPLACE(PRO.AD_NUMORIGINAL,' ',''),'-',''))
            LIKE '%{clean}%'
        OR UPPER(REPLACE(REPLACE(PRO.REFFORN,' ',''),'-',''))
            LIKE '%{clean}%'
      )
    AND ROWNUM <= 10"""

    result = await executor.execute(sql)
    cols = ["CODPROD", "PRODUTO", "MARCA", "REFERENCIA", "REFFORN",
            "AD_NUMFABRICANTE", "AD_NUMFABRICANTE2", "AD_NUMORIGINAL", "APLICACAO"]

    products = []
    if result.get("success"):
        data = result.get("data", [])
        if data and isinstance(data[0], (list, tuple)):
            rc = result.get("columns") or cols
            data = [dict(zip(rc if rc and len(rc) == len(data[0]) else cols, row)) for row in data]
        for row in data:
            if isinstance(row, dict):
                # Identificar qual campo matchou
                campo_match = "?"
                for campo in ["REFERENCIA", "AD_NUMFABRICANTE", "AD_NUMFABRICANTE2", "AD_NUMORIGINAL", "REFFORN"]:
                    val = str(row.get(campo, "") or "")
                    if clean in _sanitize_code(val):
                        campo_match = campo
                        break
                products.append({
                    "codprod": int(row.get("CODPROD", 0) or 0),
                    "produto": str(row.get("PRODUTO", "")),
                    "marca": str(row.get("MARCA", "")),
                    "referencia": str(row.get("REFERENCIA", "") or ""),
                    "aplicacao": str(row.get("APLICACAO", "") or ""),
                    "campo_match": campo_match,
                })

    print(f"[PRODUTO] resolve_manufacturer_code('{code_input}'): {len(products)} resultado(s)")
    return {"found": len(products) > 0, "products": products, "code_searched": code_input}


async def buscar_similares(codprod: int, executor) -> dict:
    """Busca codigos auxiliares/similares de um produto em AD_TGFPROAUXMMA.

    Returns:
        {"found": bool, "codprod": int, "produto": str, "marca": str,
         "auxiliares": [{"codigo": str, "marca": str, "observacao": str, "origem": str}]}
    """
    # Dados do produto
    sql_prod = f"""SELECT PRO.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA,
        PRO.REFERENCIA, PRO.AD_NUMFABRICANTE, NVL(PRO.CARACTERISTICAS,'') AS APLICACAO
    FROM TGFPRO PRO LEFT JOIN TGFMAR MAR ON MAR.CODIGO = PRO.CODMARCA
    WHERE PRO.CODPROD = {codprod}"""

    sql_aux = f"""SELECT AUX.NUMAUX AS CODIGO, NVL(MAR.DESCRICAO, 'SEM MARCA') AS MARCA,
        NVL(AUX.OBSERVACAO, '') AS OBSERVACAO, NVL(AUX.ORIGEM, '') AS ORIGEM
    FROM AD_TGFPROAUXMMA AUX
    LEFT JOIN TGFMAR MAR ON MAR.CODIGO = AUX.CODIGO
    WHERE AUX.CODPROD = {codprod}
    ORDER BY MAR.DESCRICAO, AUX.NUMAUX"""

    prod_result = await executor.execute(sql_prod)
    aux_result = await executor.execute(sql_aux)

    produto_info = {"codprod": codprod, "produto": "?", "marca": "", "referencia": "", "aplicacao": ""}
    if prod_result.get("success") and prod_result.get("data"):
        pdata = prod_result["data"]
        if pdata and isinstance(pdata[0], (list, tuple)):
            cols = prod_result.get("columns") or ["CODPROD", "PRODUTO", "MARCA", "REFERENCIA", "AD_NUMFABRICANTE", "APLICACAO"]
            pdata = [dict(zip(cols, row)) for row in pdata]
        if pdata and isinstance(pdata[0], dict):
            produto_info["produto"] = str(pdata[0].get("PRODUTO", "?"))
            produto_info["marca"] = str(pdata[0].get("MARCA", ""))
            produto_info["referencia"] = str(pdata[0].get("REFERENCIA", "") or "")
            produto_info["aplicacao"] = str(pdata[0].get("APLICACAO", "") or "")

    auxiliares = []
    if aux_result.get("success") and aux_result.get("data"):
        adata = aux_result["data"]
        if adata and isinstance(adata[0], (list, tuple)):
            cols = aux_result.get("columns") or ["CODIGO", "MARCA", "OBSERVACAO", "ORIGEM"]
            adata = [dict(zip(cols, row)) for row in adata]
        for row in adata:
            if isinstance(row, dict):
                auxiliares.append({
                    "codigo": str(row.get("CODIGO", "")),
                    "marca": str(row.get("MARCA", "")),
                    "observacao": str(row.get("OBSERVACAO", "")),
                    "origem": str(row.get("ORIGEM", "")),
                })

    print(f"[PRODUTO] buscar_similares({codprod}): {len(auxiliares)} codigo(s) auxiliar(es)")
    return {"found": len(auxiliares) > 0, **produto_info, "auxiliares": auxiliares}


async def buscar_similares_por_codigo(code_input: str, executor) -> dict:
    """Dado um codigo auxiliar, encontra o produto e lista todos os similares."""
    safe_code = _safe_sql(code_input)
    clean = _sanitize_code(safe_code)

    sql = f"""SELECT DISTINCT AUX.CODPROD
    FROM AD_TGFPROAUXMMA AUX
    WHERE UPPER(REPLACE(REPLACE(AUX.NUMAUX, ' ', ''), '-', ''))
          LIKE '%{clean}%'
    AND ROWNUM <= 5"""

    result = await executor.execute(sql)
    codprods = []
    if result.get("success") and result.get("data"):
        for row in result["data"]:
            if isinstance(row, dict):
                codprods.append(int(row.get("CODPROD", 0) or 0))
            elif isinstance(row, (list, tuple)):
                codprods.append(int(row[0] or 0))

    if not codprods:
        return {"found": False, "code_searched": code_input, "products": []}

    # Se 1 resultado, listar similares completos
    if len(codprods) == 1:
        return await buscar_similares(codprods[0], executor)

    # Se multiplos, listar os produtos encontrados
    products = []
    for cp in codprods[:5]:
        sql_p = f"SELECT PRO.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA, NVL(PRO.CARACTERISTICAS,'') AS APLICACAO FROM TGFPRO PRO LEFT JOIN TGFMAR MAR ON MAR.CODIGO=PRO.CODMARCA WHERE PRO.CODPROD={cp}"
        pr = await executor.execute(sql_p)
        if pr.get("success") and pr.get("data"):
            pdata = pr["data"]
            if pdata and isinstance(pdata[0], (list, tuple)):
                cols = pr.get("columns") or ["CODPROD", "PRODUTO", "MARCA", "APLICACAO"]
                pdata = [dict(zip(cols, row)) for row in pdata]
            if pdata and isinstance(pdata[0], dict):
                products.append({"codprod": cp, "produto": str(pdata[0].get("PRODUTO", "")), "marca": str(pdata[0].get("MARCA", "")), "aplicacao": str(pdata[0].get("APLICACAO", "") or "")})

    return {"found": True, "code_searched": code_input, "products": products, "multiple": True}


def detect_product_query(q_norm: str, params: dict) -> str | None:
    """Detecta se a pergunta e centrada em produto e qual tipo de consulta.

    Returns:
        "produto_360" - visao completa (estoque + pendencia + info)
        "busca_fabricante" - resolver codigo fabricante pra CODPROD
        "similares" - buscar cross-reference/similares
        "busca_aplicacao" - buscar por veiculo/aplicacao
        None - nao e query centrada em produto
    """
    # Busca por aplicacao: tem aplicacao mas nao codprod/codigo_fabricante
    if params.get("aplicacao") and not params.get("codprod") and not params.get("codigo_fabricante"):
        return "busca_aplicacao"

    has_product = params.get("codprod") or params.get("produto_nome") or params.get("codigo_fabricante")
    if not has_product:
        return None

    q = q_norm.lower()

    # Similares / cross-reference
    if any(w in q for w in SIMILAR_WORDS):
        return "similares"

    # Se tem codigo fabricante, precisa resolver primeiro
    if params.get("codigo_fabricante") and not params.get("codprod"):
        return "busca_fabricante"

    # Visao 360: "tudo sobre", "situacao do", "me fala tudo"
    full_view_patterns = ["tudo sobre", "situacao do", "como esta o", "me fala", "resumo do",
                          "informac", "detalhe do produto", "visao geral"]
    if any(p in q for p in full_view_patterns):
        return "produto_360"

    # Cross-intent: menciona estoque E pendencia juntos
    pend_words = {"pendente", "pendencia", "falta chegar", "pedido aberto", "compra"}
    est_words = {"estoque", "saldo", "disponivel"}
    has_pend = any(w in q for w in pend_words)
    has_est = any(w in q for w in est_words)
    if has_pend and has_est:
        return "produto_360"

    return None


def format_produto_360(prod_info: dict, estoque_data: list, pendencia_data: dict, vendas_info: dict = None) -> str:
    """Formata a visao 360 de um produto."""
    codprod = prod_info.get("codprod", "?")
    produto = prod_info.get("produto", "?")
    marca = prod_info.get("marca", "")
    ref = prod_info.get("referencia", "")

    aplicacao = prod_info.get("aplicacao", "")
    complemento = prod_info.get("complemento", "")
    num_original = prod_info.get("num_original", "")
    ref_forn = prod_info.get("ref_fornecedor", "")

    lines = []
    header = f"\U0001f4e6 **Produto {codprod} - {produto}**"
    if marca:
        header += f" ({marca})"
    lines.append(header)
    if aplicacao:
        lines.append(f"Aplicacao: {aplicacao}")
    refs = []
    if ref:
        refs.append(f"Ref: {ref}")
    if num_original and num_original != ref:
        refs.append(f"Nro. Original: {num_original}")
    if ref_forn and ref_forn != ref:
        refs.append(f"Ref. Forn: {ref_forn}")
    if refs:
        lines.append(" | ".join(refs))
    if complemento:
        lines.append(f"Complemento: {complemento}")
    lines.append("")

    # ESTOQUE
    if estoque_data:
        total_est = sum(int(float(r.get("ESTOQUE", 0) or 0)) for r in estoque_data if isinstance(r, dict))
        lines.append(f"\U0001f4ca **Estoque:** {fmt_num(total_est)} unidades\n")
        if len(estoque_data) > 1:
            lines.append("| Empresa | Estoque | Est. Min. |")
            lines.append("|---------|---------|-----------|")
            for r in estoque_data[:8]:
                if isinstance(r, dict):
                    est = fmt_num(r.get("ESTOQUE", 0))
                    estmin = fmt_num(r.get("ESTMIN", 0))
                    emp = str(r.get("EMPRESA", "?"))[:25]
                    lines.append(f"| {emp} | {est} | {estmin} |")
            lines.append("")
    else:
        lines.append("\U0001f4ca **Estoque:** sem dados\n")

    # PENDENCIA
    if pendencia_data and pendencia_data.get("detail_data"):
        detail = pendencia_data["detail_data"]
        qtd_ped = len(set(str(r.get("PEDIDO", "")) for r in detail if isinstance(r, dict)))
        vlr_total = sum(float(r.get("VLR_PENDENTE", 0) or 0) for r in detail if isinstance(r, dict))
        qtd_pend = sum(int(r.get("QTD_PENDENTE", 0) or 0) for r in detail if isinstance(r, dict))

        lines.append(f"\U0001f69a **Compras Pendentes:** {fmt_num(qtd_ped)} pedido(s), {fmt_num(qtd_pend)} un., {fmt_brl(vlr_total)}\n")
        lines.append("| Pedido | Tipo | Fornecedor | Qtd Pend. | Valor | Status |")
        lines.append("|--------|------|-----------|-----------|-------|--------|")
        shown_pedidos = set()
        for r in detail[:8]:
            if isinstance(r, dict):
                ped = str(r.get("PEDIDO", "?"))
                if ped in shown_pedidos:
                    continue
                shown_pedidos.add(ped)
                tipo = str(r.get("TIPO_COMPRA", ""))
                forn = str(r.get("FORNECEDOR", "?"))[:25]
                qtd = fmt_num(r.get("QTD_PENDENTE", 0))
                vlr = fmt_brl(r.get("VLR_PENDENTE", 0))
                status = str(r.get("STATUS_ENTREGA", "?"))
                lines.append(f"| {ped} | {tipo} | {forn} | {qtd} | {vlr} | {status} |")
        lines.append("")
    else:
        lines.append("\U0001f69a **Compras Pendentes:** nenhuma\n")

    # VENDAS
    if vendas_info and int(vendas_info.get("QTD_VENDAS", 0) or 0) > 0:
        qv = fmt_num(vendas_info.get("QTD_VENDAS", 0))
        qtdv = fmt_num(vendas_info.get("QTD_VENDIDA", 0))
        vlrv = fmt_brl(vendas_info.get("VLR_TOTAL", 0))
        lines.append(f"\U0001f4c8 **Vendas (3 meses):** {qv} notas, {qtdv} un., {vlrv}\n")

    return "\n".join(lines)


def format_busca_fabricante(resolved: dict) -> str:
    """Formata resultado da busca por codigo fabricante."""
    code = resolved.get("code_searched", "?")
    products = resolved.get("products", [])

    if not products:
        return f"\U0001f50d Nenhum produto encontrado com o codigo **{code}**.\n\nTente buscar por similares: *\"similares do {code}\"*"

    if len(products) == 1:
        p = products[0]
        response = f"\U0001f50d O codigo **{code}** corresponde ao produto:\n\n"
        response += f"| Campo | Valor |\n|---|---|\n"
        response += f"| **Codigo** | {p['codprod']} |\n"
        response += f"| **Produto** | {p['produto']} |\n"
        if p.get("marca"):
            response += f"| **Marca** | {p['marca']} |\n"
        if p.get("aplicacao"):
            response += f"| **Aplicacao** | {p['aplicacao']} |\n"
        if p.get("referencia"):
            response += f"| **Referencia** | {p['referencia']} |\n"
        response += f"| **Campo encontrado** | {p['campo_match']} |\n"
        response += f"\nQuer ver a visao completa? Pergunte: *\"tudo sobre o produto {p['codprod']}\"*"
        return response

    response = f"\U0001f50d Encontrei **{len(products)} produtos** com o codigo **{code}**:\n\n"
    response += "| CodProd | Produto | Marca | Aplicacao | Campo |\n|---------|---------|-------|-----------|-------|\n"
    for p in products:
        aplic = _trunc(p.get('aplicacao',''), 40)
        response += f"| {p['codprod']} | {str(p['produto'])[:35]} | {p.get('marca','')} | {aplic} | {p['campo_match']} |\n"
    response += f"\nEspecifique o produto pelo codigo. Ex: *\"tudo sobre o produto {products[0]['codprod']}\"*"
    return response


def format_similares(sim_data: dict) -> str:
    """Formata resultado da busca de similares/cross-reference."""
    if not sim_data.get("found"):
        code = sim_data.get("code_searched", sim_data.get("codprod", "?"))
        return f"Nao encontrei codigos auxiliares/similares para **{code}**."

    # Se veio de busca por codigo texto com multiplos produtos
    if sim_data.get("multiple"):
        products = sim_data.get("products", [])
        response = f"\U0001f504 O codigo **{sim_data.get('code_searched', '?')}** aparece em {len(products)} produtos:\n\n"
        response += "| CodProd | Produto | Marca |\n|---------|---------|-------|\n"
        for p in products:
            response += f"| {p['codprod']} | {str(p['produto'])[:40]} | {p.get('marca','')} |\n"
        response += f"\nPara ver similares de um produto especifico: *\"similares do produto {products[0]['codprod']}\"*"
        return response

    codprod = sim_data.get("codprod", "?")
    produto = sim_data.get("produto", "?")
    marca = sim_data.get("marca", "")
    aplicacao = sim_data.get("aplicacao", "")
    auxiliares = sim_data.get("auxiliares", [])

    response = f"\U0001f504 **Similares do produto {codprod} - {produto}**"
    if marca:
        response += f" ({marca})"
    if aplicacao:
        response += f"\nAplicacao: {aplicacao}"
    response += f"\n\nEncontrei **{len(auxiliares)}** codigo(s) auxiliar(es):\n\n"

    # Agrupar por marca
    from collections import defaultdict
    por_marca = defaultdict(list)
    for aux in auxiliares:
        por_marca[aux.get("marca", "?")].append(aux)

    response += "| Codigo | Marca | Obs. |\n|--------|-------|------|\n"
    count = 0
    for m_name in sorted(por_marca.keys()):
        for aux in por_marca[m_name][:5]:
            obs = str(aux.get("observacao", ""))[:20]
            response += f"| {aux['codigo']} | {m_name} | {obs} |\n"
            count += 1
            if count >= 30:
                break
        if count >= 30:
            break

    if len(auxiliares) > 30:
        response += f"\n*...e mais {len(auxiliares) - 30} codigo(s).*"

    response += f"\n\n{len(por_marca)} marca(s) diferente(s)."
    return response


# ============================================================
# VIEWS AGREGADAS - Perguntas "quem compra/fornece marca X?"
# ============================================================

def detect_aggregation_view(question_norm: str) -> str | None:
    """Detecta se a pergunta pede uma visao agregada em vez de listagem.

    Returns:
        "comprador_marca" - quem compra marca X
        "fornecedor_marca" - quem fornece marca X
        None - pergunta normal (listagem)
    """
    q = question_norm.lower()

    # "quem compra/e o comprador/e responsavel pela marca X"
    if re.search(r'quem\s+(compra|e\s+o?\s*comprador|e\s+responsavel)', q):
        return "comprador_marca"
    if re.search(r'comprador(es?)?\s+(da|de|do)\s+', q):
        return "comprador_marca"
    if re.search(r'responsavel\s+(pela|pela\s+marca|pelas?\s+compras?\s+d)', q):
        return "comprador_marca"

    # "quem fornece/e o fornecedor da marca X"
    if re.search(r'quem\s+(fornece|e\s+o?\s*fornecedor|vende|entrega)', q):
        return "fornecedor_marca"
    if re.search(r'fornecedor(es?)?\s+(da|de|do)\s+marca', q):
        return "fornecedor_marca"

    # "fornecedor da SABO" sem "pedido/pendencia" = quem fornece
    if re.search(r'fornecedor(es?)?\s+(da|de|do)\s+\w+', q) and not re.search(r'(pedido|pendencia|pend)', q):
        return "fornecedor_marca"

    return None


def format_comprador_marca(detail_data: list, marca: str) -> str:
    """Formata resposta 'quem compra marca X' agrupando por COMPRADOR."""
    from collections import defaultdict
    compradores = defaultdict(lambda: {"pedidos": set(), "itens": 0, "valor": 0.0})

    for r in detail_data:
        if not isinstance(r, dict):
            continue
        comp = r.get("COMPRADOR") or "SEM COMPRADOR"
        compradores[comp]["pedidos"].add(str(r.get("PEDIDO", "")))
        compradores[comp]["itens"] += 1
        compradores[comp]["valor"] += float(r.get("VLR_PENDENTE", 0) or 0)

    if not compradores:
        return f"Nao encontrei pedidos pendentes da marca {marca}."

    sorted_comp = sorted(compradores.items(), key=lambda x: -x[1]["valor"])

    response = f"\U0001f3f7\ufe0f **Comprador(es) da marca {marca}:**\n\n"
    response += "| Comprador | Pedidos | Itens | Valor Pendente |\n|---|---|---|---|\n"
    for comp, data in sorted_comp:
        response += f"| {comp} | {len(data['pedidos'])} | {data['itens']} | R$ {fmt_num(data['valor'])} |\n"

    if len(sorted_comp) == 1:
        response += f"\n**{sorted_comp[0][0]}** e o comprador responsavel pela marca {marca}."

    return response


def format_fornecedor_marca(detail_data: list, marca: str) -> str:
    """Formata resposta 'quem fornece marca X' agrupando por FORNECEDOR."""
    from collections import defaultdict
    fornecedores = defaultdict(lambda: {"pedidos": set(), "itens": 0, "valor": 0.0})

    for r in detail_data:
        if not isinstance(r, dict):
            continue
        forn = r.get("FORNECEDOR") or "?"
        fornecedores[forn]["pedidos"].add(str(r.get("PEDIDO", "")))
        fornecedores[forn]["itens"] += 1
        fornecedores[forn]["valor"] += float(r.get("VLR_PENDENTE", 0) or 0)

    if not fornecedores:
        return f"Nao encontrei fornecedores com pedidos pendentes da marca {marca}."

    sorted_forn = sorted(fornecedores.items(), key=lambda x: -x[1]["valor"])

    response = f"\U0001f3ed **Fornecedor(es) da marca {marca}:**\n\n"
    response += "| Fornecedor | Pedidos | Itens | Valor Pendente |\n|---|---|---|---|\n"
    for forn, data in sorted_forn:
        response += f"| {forn} | {len(data['pedidos'])} | {data['itens']} | R$ {fmt_num(data['valor'])} |\n"

    return response


def format_pendencia_response(kpis_data, detail_data, description, params, view_mode="pedidos", extra_columns=None):
    if not kpis_data:
        filtro = params.get("marca") or params.get("fornecedor") or params.get("empresa") or ""
        return f"Nao encontrei pedidos pendentes{' para ' + filtro if filtro else ''}. Verifique se o nome esta correto."

    row = kpis_data[0] if isinstance(kpis_data[0], dict) else {}
    qtd_ped = int(row.get("QTD_PEDIDOS", 0) or 0)
    qtd_itens = int(row.get("QTD_ITENS", 0) or 0)
    vlr = float(row.get("VLR_PENDENTE", 0) or 0)

    if qtd_ped == 0:
        filtro = params.get("marca") or params.get("fornecedor") or ""
        return f"Nao encontrei pedidos pendentes{' para ' + filtro if filtro else ''}."

    lines = []
    lines.append(f"\U0001f4e6 **{description.title()}**\n")
    s_ped = "s" if qtd_ped > 1 else ""
    s_it = "ns" if qtd_itens > 1 else "m"
    lines.append(f"Voce tem **{fmt_num(qtd_ped)} pedido{s_ped}** pendente{s_ped}, com **{fmt_num(qtd_itens)} ite{s_it}** e valor total de **{fmt_brl(vlr)}**.\n")

    if detail_data:
        if view_mode == "itens":
            # Colunas base
            has_aplic = any(item.get("APLICACAO") for item in detail_data[:12] if isinstance(item, dict))
            base_cols = ["PEDIDO", "CODPROD", "PRODUTO"]
            base_cols.append("APLICACAO" if has_aplic else "MARCA")

            # Inserir extras ANTES das colunas numericas
            if extra_columns:
                for ec in extra_columns:
                    if ec not in base_cols:
                        base_cols.append(ec)

            base_cols.extend(["QTD_PENDENTE", "VLR_PENDENTE", "STATUS_ENTREGA"])
            visible_cols = base_cols

            # Mensagem de colunas extras
            if extra_columns:
                added_labels = [COLUMN_LABELS.get(c, c) for c in extra_columns]
                lines.append(f"\u2705 Coluna{'s' if len(added_labels)>1 else ''} extra{'s' if len(added_labels)>1 else ''}: **{', '.join(added_labels)}**\n")

            # Header
            headers = [COLUMN_LABELS.get(c, c) for c in visible_cols]
            lines.append("**Itens pendentes:**\n")
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("|" + "|".join(["---" for _ in visible_cols]) + "|")

            # Rows
            for item in detail_data[:12]:
                if not isinstance(item, dict):
                    continue
                cells = []
                for c in visible_cols:
                    val = item.get(c, "")
                    if val is None: val = ""
                    val = str(val)
                    max_w = COLUMN_MAX_WIDTH.get(c, 40)
                    if len(val) > max_w:
                        val = val[:max_w-1] + "\u2026"
                    if "VLR" in c:
                        try: val = fmt_brl(float(val))
                        except: pass
                    elif c in ("QTD_PENDENTE", "QTD_PEDIDA", "QTD_ATENDIDA", "DIAS_ABERTO"):
                        try: val = str(int(float(val or 0)))
                        except: pass
                    cells.append(val)
                lines.append("| " + " | ".join(cells) + " |")

            if len(detail_data) > 12:
                lines.append(f"\n*...e mais {len(detail_data) - 12} itens.*\n")
        else:
            pedidos = {}
            for item in detail_data:
                if not isinstance(item, dict): continue
                ped = item.get("PEDIDO", "?")
                if ped not in pedidos:
                    pedidos[ped] = {
                        "PEDIDO": ped,
                        "FORNECEDOR": str(item.get("FORNECEDOR",""))[:30],
                        "DT_PEDIDO": item.get("DT_PEDIDO",""),
                        "STATUS_ENTREGA": item.get("STATUS_ENTREGA","?"),
                        "_itens": 0, "_valor": 0.0
                    }
                    if extra_columns:
                        for ec in extra_columns:
                            pedidos[ped][ec] = str(item.get(ec, "") or "")
                pedidos[ped]["_itens"] += 1
                pedidos[ped]["_valor"] += float(item.get("VLR_PENDENTE", 0) or 0)

            # Colunas base
            base_cols = ["PEDIDO", "FORNECEDOR", "DT_PEDIDO"]
            if extra_columns:
                for ec in extra_columns:
                    if ec not in base_cols:
                        base_cols.append(ec)
            base_cols.extend(["_itens", "_valor", "STATUS_ENTREGA"])

            label_override = {"_itens": "Itens", "_valor": "Valor Pendente"}
            headers = [label_override.get(c, COLUMN_LABELS.get(c, c)) for c in base_cols]

            if extra_columns:
                added_labels = [COLUMN_LABELS.get(c, c) for c in extra_columns]
                lines.append(f"\u2705 Coluna{'s' if len(added_labels)>1 else ''} extra{'s' if len(added_labels)>1 else ''}: **{', '.join(added_labels)}**\n")

            lines.append("**Pedidos:**\n")
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("|" + "|".join(["---" for _ in base_cols]) + "|")

            for pd in list(pedidos.values())[:10]:
                cells = []
                for c in base_cols:
                    if c == "_valor": cells.append(fmt_brl(pd["_valor"]))
                    elif c == "_itens": cells.append(str(pd["_itens"]))
                    else:
                        val = str(pd.get(c, ""))
                        max_w = COLUMN_MAX_WIDTH.get(c, 40)
                        if len(val) > max_w: val = val[:max_w-1] + "\u2026"
                        cells.append(val)
                lines.append("| " + " | ".join(cells) + " |")

            if len(pedidos) > 10:
                lines.append(f"\n*...e mais {len(pedidos) - 10} pedidos.*\n")

    lines.append(f"\n\U0001f4e5 **Quer que eu gere um arquivo Excel com todos os {fmt_num(qtd_itens)} itens?**")
    return "\n".join(lines)


def format_vendas_response(kpis_data, periodo_nome):
    qtd = int(kpis_data.get("QTD_VENDAS", 0) or 0)
    fat = float(kpis_data.get("FATURAMENTO", 0) or 0)
    ticket = float(kpis_data.get("TICKET_MEDIO", 0) or 0)
    if qtd == 0:
        return f"Nao encontrei vendas para o periodo **{periodo_nome}**."
    lines = [f"\U0001f4ca **Vendas - {periodo_nome.title()}**\n",
             f"**{fmt_num(qtd)}** notas de venda com faturamento total de **{fmt_brl(fat)}**.",
             f"Ticket medio: **{fmt_brl(ticket)}**.\n"]
    return "\n".join(lines)


def format_estoque_response(data, params):
    if not data:
        filtro = params.get("codprod") or params.get("produto_nome") or params.get("marca") or ""
        return f"Nao encontrei informacoes de estoque{' para ' + str(filtro) if filtro else ''}."
    lines = []
    if params.get("codprod") or params.get("produto_nome"):
        row = data[0] if isinstance(data[0], dict) else {}
        lines.append(f"\U0001f4e6 **Estoque do produto {row.get('CODPROD','?')}**\n")
        lines.append(f"**{row.get('PRODUTO','?')}**" + (f" ({row.get('MARCA','')})" if row.get('MARCA') else ""))
        if row.get('APLICACAO'):
            lines.append(f"Aplicacao: {row.get('APLICACAO','')}")
        lines.append(f"Estoque atual: **{fmt_num(row.get('ESTOQUE',0))}** unidades")
        if row.get('ESTMIN'):
            lines.append(f"Estoque minimo: **{fmt_num(row.get('ESTMIN',0))}**")
            if int(float(row.get('ESTOQUE',0) or 0)) <= int(float(row.get('ESTMIN',0) or 0)):
                lines.append("\u26a0\ufe0f **Abaixo do estoque minimo!**")
        if len(data) > 1:
            lines.append("\n**Por empresa:**\n| Empresa | Estoque | Est. Minimo |\n|---------|---------|-------------|")
            for r in data[:10]:
                if isinstance(r, dict):
                    lines.append(f"| {str(r.get('EMPRESA','?'))[:25]} | {fmt_num(r.get('ESTOQUE',0))} | {fmt_num(r.get('ESTMIN',0))} |")
    else:
        lines.append(f"\u26a0\ufe0f **Estoque Critico** - {len(data)} produto{'s' if len(data)>1 else ''}\n")
        lines.append("| CodProd | Produto | Marca | Estoque | Est. Min. |\n|---------|---------|-------|---------|-----------|")
        for r in data[:15]:
            if isinstance(r, dict):
                lines.append(f"| {r.get('CODPROD','?')} | {str(r.get('PRODUTO',''))[:30]} | {str(r.get('MARCA',''))[:15]} | {fmt_num(r.get('ESTOQUE',0))} | {fmt_num(r.get('ESTMIN',0))} |")
        if len(data) > 15:
            lines.append(f"\n*...e mais {len(data)-15} produtos.*")
    return "\n".join(lines)


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(data, columns, filename, title=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return generate_csv(data, columns, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    hfill = PatternFill(start_color="0E75B9", end_color="0E75B9", fill_type="solid")
    hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    dfont = Font(name="Arial", size=9)
    brd = Border(bottom=Side(style="thin", color="E0E0E0"))
    sr = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.cell(row=1, column=1, value=title).font = Font(name="Arial", size=12, bold=True, color="0E75B9")
        ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=8, color="888888")
        sr = 4
    for ci, cn in enumerate(columns, 1):
        c = ws.cell(row=sr, column=ci, value=cn); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
    ccols = {"VLR_UNITARIO","VLR_PENDENTE","VLR_TOTAL_PENDENTE","VALOR","FATURAMENTO","TICKET_MEDIO"}
    ncols = {"QTD_PEDIDA","QTD_ATENDIDA","QTD_PENDENTE","DIAS_ABERTO","CODPROD","PEDIDO","ESTOQUE","ESTMIN","QTD"}
    for ri, rd in enumerate(data, sr + 1):
        for ci, cn in enumerate(columns, 1):
            val = rd.get(cn, "") if isinstance(rd, dict) else (rd[ci-1] if ci <= len(rd) else "")
            cell = ws.cell(row=ri, column=ci, value=val); cell.font = dfont; cell.border = brd
            if cn in ccols:
                try: cell.value = float(val or 0); cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                except: pass
            elif cn in ncols:
                try: cell.value = int(float(val or 0)); cell.alignment = Alignment(horizontal="right")
                except: pass
    for ci, cn in enumerate(columns, 1):
        ml = len(str(cn))
        for ri in range(sr+1, min(sr+50, len(data)+sr+1)):
            cv = ws.cell(row=ri, column=ci).value
            if cv: ml = max(ml, min(len(str(cv)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = ml + 3
    ws.auto_filter.ref = f"A{sr}:{get_column_letter(len(columns))}{sr + len(data)}"
    ws.freeze_panes = ws.cell(row=sr + 1, column=1)
    static_dir = Path(__file__).parent.parent / "api" / "static" / "exports"
    static_dir.mkdir(parents=True, exist_ok=True)
    fp = static_dir / filename; wb.save(str(fp)); return str(fp)


def generate_csv(data, columns, filename):
    static_dir = Path(__file__).parent.parent / "api" / "static" / "exports"
    static_dir.mkdir(parents=True, exist_ok=True)
    fp = static_dir / filename.replace(".xlsx", ".csv")
    with open(fp, "w", encoding="utf-8-sig") as f:
        f.write(";".join(columns) + "\n")
        for row in data:
            vals = [str(row.get(c,"")).replace(";",",") for c in columns] if isinstance(row, dict) else [str(v).replace(";",",") for v in row]
            f.write(";".join(vals) + "\n")
    return str(fp)


# ============================================================
# CONVERSATION CONTEXT - Memoria por usuario
# ============================================================

# Palavras que indicam referencia ao contexto anterior
FOLLOWUP_WORDS = {
    "desses", "destes", "daqueles", "delas", "deles",
    "esses", "estes", "aqueles", "essas", "estas", "aquelas",
    "neles", "nelas", "nisso", "nesse", "nessa", "neste", "nesta",
    "mesma", "mesmo", "mesmos", "mesmas",
    "tambem", "alem", "ainda", "mais",
    "agora", "entao",
}

# Patterns que indicam follow-up (referencia a dados anteriores)
FOLLOWUP_PATTERNS = [
    r'\b(desse[s]?|deste[s]?|daquela?[s]?|dela[s]?|dele[s]?)\b',
    r'\b(esse[s]?|este[s]?|aquele[s]?|essa[s]?|esta[s]?|aquela[s]?)\b',
    r'\b(nele[s]?|nela[s]?|nisso|nesse[s]?|nessa[s]?)\b',
    r'\be (os|as|a|o) (itens|pedidos|pendentes|atrasados)\b',
    r'^e\s+(os|as|quais|quantos|quantas|qual|quanto)\b',
    r'^e\s+(da|do|de|das|dos)\s+',
    r'^(quais|quantos|quantas|qual|quanto)\s+(sao|estao|tem)\s+(os|as|atrasad)',
]

# Patterns que indicam filtro sobre dados anteriores
# Regras de filtro/ordenacao sobre dados anteriores
# ORDEM IMPORTA: patterns mais longos/especificos primeiro
FILTER_RULES = [
    # === TIPO DE COMPRA ===
    {"match": ["compra casada", "compras casadas", "pedido casado", "pedidos casados", "empenho", "empenhado", "empenhados", "vinculada"],
     "filter": {"TIPO_COMPRA": "Casada"}},
    {"match": ["compra estoque", "compra de estoque", "compra para estoque", "compras de estoque", "entrega futura", "reposicao"],
     "filter": {"TIPO_COMPRA": "Estoque"}},
    # === PREVISAO DE ENTREGA (mais especificos primeiro!) ===
    {"match": ["maior data de entrega", "maior previsao de entrega", "maior previsao entrega"],
     "sort": "PREVISAO_ENTREGA_DESC", "top": 1},
    {"match": ["menor data de entrega", "menor previsao de entrega", "menor previsao entrega"],
     "sort": "PREVISAO_ENTREGA_ASC", "top": 1},
    {"match": ["data de entrega mais distante", "previsao mais distante", "entrega mais longe"],
     "sort": "PREVISAO_ENTREGA_DESC", "top": 1},
    {"match": ["data de entrega mais proxima", "previsao mais proxima", "proxima entrega"],
     "sort": "PREVISAO_ENTREGA_ASC", "top": 1},
    # === SUPERLATIVOS (sort + top N) ===
    {"match": ["mais atrasado"],     "sort": "DIAS_ABERTO_DESC",      "top": 1},
    {"match": ["mais atrasados"],    "sort": "DIAS_ABERTO_DESC",      "top": 5},
    {"match": ["mais caro"],         "sort": "VLR_PENDENTE_DESC",     "top": 1},
    {"match": ["mais caros"],        "sort": "VLR_PENDENTE_DESC",     "top": 5},
    {"match": ["mais barato"],       "sort": "VLR_PENDENTE_ASC",      "top": 1},
    {"match": ["mais baratos"],      "sort": "VLR_PENDENTE_ASC",      "top": 5},
    {"match": ["mais antigo"],       "sort": "DIAS_ABERTO_DESC",      "top": 1},
    {"match": ["mais antigos"],      "sort": "DIAS_ABERTO_DESC",      "top": 5},
    {"match": ["mais recente"],      "sort": "DT_PEDIDO_DESC",        "top": 1},
    {"match": ["mais recentes"],     "sort": "DT_PEDIDO_DESC",        "top": 5},
    {"match": ["maior valor"],       "sort": "VLR_PENDENTE_DESC",     "top": 1},
    {"match": ["menor valor"],       "sort": "VLR_PENDENTE_ASC",      "top": 1},
    {"match": ["maior quantidade"],  "sort": "QTD_PENDENTE_DESC",     "top": 1},
    {"match": ["mais urgente"],      "sort": "DIAS_ABERTO_DESC",      "top": 1},
    {"match": ["mais urgentes"],     "sort": "DIAS_ABERTO_DESC",      "top": 5},
    # === FILTROS por CAMPO (mais especificos primeiro) ===
    {"match": ["sem previsao de entrega", "sem data de entrega", "sem previsao entrega"],
     "filter_fn": "empty", "filter_field": "PREVISAO_ENTREGA"},
    {"match": ["sem confirmacao", "nao confirmado", "nao confirmados"],
     "filter": {"CONFIRMADO": "N"}},
    {"match": ["confirmado", "confirmados"],           "filter": {"CONFIRMADO": "S"}},
    # === FILTROS por STATUS_ENTREGA ===
    {"match": ["sem previsao"],                        "filter": {"STATUS_ENTREGA": "SEM PREVISAO"}},
    {"match": ["no prazo", "dentro do prazo"],         "filter": {"STATUS_ENTREGA": "NO PRAZO"}},
    {"match": ["atrasado", "atrasados"],               "filter": {"STATUS_ENTREGA": "ATRASADO"}},
    {"match": ["proximo", "proximos"],                 "filter": {"STATUS_ENTREGA": "PROXIMO"}},
]


# ============================================================
# COMPILED KNOWLEDGE (auto-gerado pelo Knowledge Compiler)
# Merge com manual: manual SEMPRE ganha, compilado complementa
# ============================================================

_COMPILED_SCORES = {}   # {intent: {word: weight}}
_COMPILED_RULES = []    # [{match: [...], filter/sort/top: ...}]
_COMPILED_EXAMPLES = [] # Groq examples extras
_COMPILED_SYNONYMS = [] # Sinonimos extras
_COMPILED_LOADED = False


def _load_compiled_knowledge():
    """Carrega compiled_knowledge.json e monta dicts de merge."""
    global _COMPILED_SCORES, _COMPILED_RULES, _COMPILED_EXAMPLES, _COMPILED_SYNONYMS, _COMPILED_LOADED
    if _COMPILED_LOADED:
        return
    _COMPILED_LOADED = True

    if not COMPILED_PATH.exists():
        return

    try:
        with open(COMPILED_PATH, "r", encoding="utf-8") as f:
            compiled = json.load(f)
    except Exception as e:
        print(f"[SMART] Erro ao carregar compiled_knowledge: {e}")
        return

    # 1. Keywords -> _COMPILED_SCORES (formato: {intent: {word: weight}})
    for intent, keywords in compiled.get("intent_keywords", {}).items():
        if intent == "unknown":
            continue
        if intent not in _COMPILED_SCORES:
            _COMPILED_SCORES[intent] = {}
        for kw in keywords:
            word = kw.get("word", "").lower().strip()
            weight = kw.get("weight", 3)
            if word and len(word) >= 2:
                # Se ja existe no manual, NAO sobrescrever
                if intent in INTENT_SCORES and word in INTENT_SCORES[intent]:
                    continue
                _COMPILED_SCORES[intent][word] = weight

    # 2. Filter rules -> _COMPILED_RULES
    for rule in compiled.get("filter_rules", []):
        matches = rule.get("match", [])
        if not matches:
            continue
        # Verificar se ja existe no manual
        manual_matches = set()
        for mr in FILTER_RULES:
            for m in mr.get("match", []):
                manual_matches.add(m.lower())
        if all(m.lower() in manual_matches for m in matches):
            continue
        _COMPILED_RULES.append(rule)

    # 3. Groq examples
    _COMPILED_EXAMPLES.extend(compiled.get("groq_examples", []))

    # 4. Synonyms
    _COMPILED_SYNONYMS.extend(compiled.get("synonyms", []))

    total_kw = sum(len(v) for v in _COMPILED_SCORES.values())
    print(f"[SMART] Compiled knowledge: +{total_kw} keywords em {len(_COMPILED_SCORES)} intents, "
          f"+{len(_COMPILED_RULES)} filter rules, +{len(_COMPILED_EXAMPLES)} examples")

    # Intents potenciais
    for pi in compiled.get("potential_intents", []):
        print(f"[SMART] ** Intent potencial: {pi['name']} ({pi['keywords_count']} keywords) - {pi.get('note', '')}")


# ============================================================
# COLUNAS DINAMICAS - Personalizacao de relatorios via Groq
# ============================================================

# Campos que JA existem no SQL detail (nao precisa modificar query)
EXISTING_SQL_COLUMNS = {
    "EMPRESA", "PEDIDO", "TIPO_COMPRA", "COMPRADOR", "DT_PEDIDO",
    "PREVISAO_ENTREGA", "CONFIRMADO", "FORNECEDOR", "CODPROD", "PRODUTO",
    "MARCA", "APLICACAO", "UNIDADE", "QTD_PEDIDA", "QTD_ATENDIDA",
    "QTD_PENDENTE", "VLR_UNITARIO", "VLR_PENDENTE", "DIAS_ABERTO", "STATUS_ENTREGA",
    "NUM_FABRICANTE",
}

# Campos EXTRAS que precisam ser adicionados ao SQL
EXTRA_SQL_FIELDS = {
    "NUM_ORIGINAL":    "NVL(PRO.AD_NUMORIGINAL, '') AS NUM_ORIGINAL",
    "REFERENCIA":      "NVL(PRO.REFERENCIA, '') AS REFERENCIA",
    "REF_FORNECEDOR":  "NVL(PRO.REFFORN, '') AS REF_FORNECEDOR",
    "COMPLEMENTO":     "NVL(PRO.COMPLDESC, '') AS COMPLEMENTO",
    "NCM":             "NVL(PRO.NCM, '') AS NCM",
}

# Normalizacao (Groq pode retornar variacoes)
COLUMN_NORMALIZE = {
    "PREVISAO": "PREVISAO_ENTREGA", "PREVISAO_ENTREGA": "PREVISAO_ENTREGA",
    "DIAS": "DIAS_ABERTO", "DIAS_ABERTO": "DIAS_ABERTO",
    "FABRICANTE": "NUM_FABRICANTE", "NUM_FABRICANTE": "NUM_FABRICANTE",
    "NUMERO_FABRICANTE": "NUM_FABRICANTE", "CODIGO_FABRICANTE": "NUM_FABRICANTE",
    "ORIGINAL": "NUM_ORIGINAL", "NUM_ORIGINAL": "NUM_ORIGINAL",
    "REFERENCIA": "REFERENCIA", "EMPRESA": "EMPRESA",
    "TIPO_COMPRA": "TIPO_COMPRA", "COMPRADOR": "COMPRADOR",
    "FORNECEDOR": "FORNECEDOR", "CONFIRMADO": "CONFIRMADO",
    "UNIDADE": "UNIDADE", "QTD_PEDIDA": "QTD_PEDIDA",
    "QTD_ATENDIDA": "QTD_ATENDIDA", "VLR_UNITARIO": "VLR_UNITARIO",
    "APLICACAO": "APLICACAO",
}

# Labels amigaveis para headers de tabela
COLUMN_LABELS = {
    "PEDIDO": "Pedido", "CODPROD": "CodProd", "PRODUTO": "Produto",
    "MARCA": "Marca", "QTD_PENDENTE": "Qtd Pend.", "VLR_PENDENTE": "Valor",
    "STATUS_ENTREGA": "Status", "FORNECEDOR": "Fornecedor", "DT_PEDIDO": "Data",
    "PREVISAO_ENTREGA": "Previsao", "EMPRESA": "Empresa", "TIPO_COMPRA": "Tipo",
    "COMPRADOR": "Comprador", "CONFIRMADO": "Confirmado", "APLICACAO": "Aplicacao",
    "UNIDADE": "Unid.", "QTD_PEDIDA": "Qtd Pedida", "QTD_ATENDIDA": "Qtd Atend.",
    "VLR_UNITARIO": "Vlr Unit.", "DIAS_ABERTO": "Dias",
    "NUM_FABRICANTE": "Cod. Fabricante", "NUM_ORIGINAL": "Nro. Original",
    "REFERENCIA": "Referencia", "REF_FORNECEDOR": "Ref. Forn.",
    "COMPLEMENTO": "Complemento", "NCM": "NCM",
}

COLUMN_MAX_WIDTH = {
    "PRODUTO": 30, "FORNECEDOR": 25, "EMPRESA": 20, "APLICACAO": 25,
    "COMPLEMENTO": 25, "NUM_FABRICANTE": 18, "REFERENCIA": 18, "COMPRADOR": 15,
}


def _is_complex_query(q_norm: str, tokens: list, pattern_filters: dict) -> bool:
    """Detecta se a query tem complexidade alem de intent+entidade simples.
    Se sim, vale chamar Groq pra interpretar mesmo quando scoring resolveu o intent."""
    # Se pattern ja resolveu filtros, nao precisa Groq
    if pattern_filters:
        return False

    # Palavras que indicam query complexa (filtros, ordenacao, comparacao)
    complexity_words = {
        "maior", "menor", "mais", "menos", "primeiro", "ultimo", "ultima",
        "acima", "abaixo", "entre", "superior", "inferior",
        "sem", "com",
        "previsao", "entrega", "confirmado", "confirmacao",
        "prazo", "atraso", "atrasado", "atrasados",
        "caro", "barato", "urgente", "critico",
        "valor", "quantidade", "data",
        "diferente", "igual", "mesmo", "vazio",
    }
    found = [t for t in tokens if t in complexity_words]
    if len(found) >= 2:
        return True

    # Patterns explicitos de complexidade
    complex_patterns = [
        r'(?:maior|menor|mais|menos)\s+(?:data|valor|quantidade|previsao|prazo|dias)',
        r'(?:sem|com)\s+(?:previsao|confirmacao|data|entrega)',
        r'(?:acima|abaixo)\s+de\s+\d',
        r'(?:entre)\s+\d.*e\s+\d',
        r'(?:data|previsao)\s+de\s+entrega',
    ]
    for p in complex_patterns:
        if re.search(p, q_norm):
            return True

    return False


def _llm_to_filters(llm_result: dict, question: str = "") -> dict:
    """Converte o resultado da LLM (filtro/ordenar/top) no formato de apply_filters.
    Inclui pos-processamento pra corrigir confusoes comuns da LLM."""
    filters = {}

    # Filtro estruturado da LLM
    filtro = llm_result.get("filtro")
    if filtro and isinstance(filtro, dict):
        campo = filtro.get("campo", "")
        operador = filtro.get("operador", "")
        valor = filtro.get("valor")

        if operador == "igual" and campo and valor:
            filters[campo] = str(valor).upper()
        elif operador == "vazio" and campo:
            filters[f"_fn_empty"] = campo
        elif operador == "nao_vazio" and campo:
            filters[f"_fn_not_empty"] = campo
        elif operador == "maior" and campo and valor:
            filters[f"_fn_maior"] = f"{campo}:{valor}"
        elif operador == "menor" and campo and valor:
            filters[f"_fn_menor"] = f"{campo}:{valor}"
        elif operador == "contem" and campo and valor:
            filters[f"_fn_contem"] = f"{campo}:{valor}"

    # Ordenacao
    ordenar = llm_result.get("ordenar")
    if ordenar and isinstance(ordenar, str):
        filters["_sort"] = ordenar.upper()

    # Top N
    top = llm_result.get("top")
    if top and isinstance(top, (int, float)):
        filters["_top"] = int(top)
    elif top and isinstance(top, str) and top.isdigit():
        filters["_top"] = int(top)

    # tipo_compra (campo extra retornado pelo Groq)
    tipo_compra = llm_result.get("tipo_compra")
    if tipo_compra and isinstance(tipo_compra, str):
        tc = tipo_compra.lower().strip()
        if tc in ("casada", "empenho", "vinculada"):
            filters["TIPO_COMPRA"] = "Casada"
        elif tc in ("estoque", "futura", "reposicao"):
            filters["TIPO_COMPRA"] = "Estoque"

    # === POS-PROCESSAMENTO: corrigir confusoes comuns da LLM ===
    if question:
        q_lower = question.lower()
        sort_key = filters.get("_sort", "")

        # Confusao #1: "data de entrega" / "previsao de entrega" → LLM retorna DT_PEDIDO
        entrega_words = ["entrega", "previsao", "chegar", "chegada", "previsão"]
        pedido_words = ["data do pedido", "quando pediu", "quando comprou", "data da compra"]
        mentions_entrega = any(w in q_lower for w in entrega_words)
        mentions_pedido = any(w in q_lower for w in pedido_words)

        if mentions_entrega and not mentions_pedido:
            if "DT_PEDIDO" in sort_key:
                old_sort = sort_key
                filters["_sort"] = sort_key.replace("DT_PEDIDO", "PREVISAO_ENTREGA")
                print(f"[LLM-FIX] Sort corrigido: {old_sort} -> {filters['_sort']} (query menciona entrega)")

    return filters


def detect_followup(tokens: list, question_norm: str) -> bool:
    """Detecta se a pergunta e um follow-up referenciando dados anteriores."""
    # Check tokens (pronomes, advs de continuidade)
    if any(t in FOLLOWUP_WORDS for t in tokens):
        return True
    # Check regex patterns
    for pattern in FOLLOWUP_PATTERNS:
        if re.search(pattern, question_norm):
            return True
    # Pergunta curta com palavras de follow-up implicito
    # Ex: "quantos atrasados?", "qual o mais caro?"
    # Mas NAO pra queries completas como "pedidos pendentes da Donaldson"
    if len(tokens) <= 7:
        followup_indicators = {"itens", "pedidos", "atrasados", "atrasado", "pendentes",
                                "pendente", "confirmados", "prazo", "previsao", "proximo",
                                "proximos", "urgente", "urgentes", "caro", "caros",
                                "barato", "baratos", "antigo", "antigos", "recente", "recentes"}
        has_indicator = any(t in followup_indicators for t in tokens)
        has_qualifier = any(t in tokens for t in ["marca", "fornecedor", "empresa", "produto"])
        if has_indicator and not has_qualifier:
            noise_words = {"quais", "quantos", "quantas", "estao", "sao", "como", "qual", "quem",
                           "itens", "pedidos", "atrasados", "atrasado", "pendentes", "pendente",
                           "confirmados", "valor", "maior", "menor", "mais", "menos",
                           "prazo", "previsao", "proximo", "proximos", "urgente", "caro",
                           "barato", "antigo", "recente", "total", "todos", "todas",
                           "pedido", "esta", "esse", "essa", "qual"}
            other_words = [t for t in tokens if t not in noise_words and len(t) >= 4]
            if not other_words:
                return True
    return False


def detect_filter_request(question_norm: str, tokens: list) -> dict:
    """Detecta se o usuario quer filtrar/ordenar dados anteriores. FILTER_RULES + compiladas."""
    result = {}

    # Manual primeiro, depois compiladas
    all_rules = list(FILTER_RULES) + list(_COMPILED_RULES)
    for rule in all_rules:
        matched = any(m in question_norm for m in rule["match"])
        if not matched:
            continue

        if "filter" in rule:
            result.update(rule["filter"])
        if "filter_fn" in rule:
            # Filtro especial (ex: campo vazio)
            result[f"_fn_{rule['filter_fn']}"] = rule["filter_field"]
        if "sort" in rule:
            if "_sort" not in result:  # Primeiro sort ganha
                result["_sort"] = rule["sort"]
        if "top" in rule:
            if "_top" not in result:  # Primeiro top ganha
                result["_top"] = rule["top"]

        # Se regra tem sort/top, parar (evitar conflito de ordenacao)
        # Se regra so tem filter, continuar acumulando (ex: TIPO_COMPRA + STATUS_ENTREGA)
        if "sort" in rule or "top" in rule:
            break

    # Detectar numero explicito na pergunta: "5 mais caros", "top 10", "3 primeiros"
    num_match = re.search(r'\b(\d{1,3})\s+(?:mais|primeiro|primeiros|maior|menor|ultim)', question_norm)
    if not num_match:
        num_match = re.search(r'(?:top|os)\s+(\d{1,3})\b', question_norm)
    if num_match:
        result["_top"] = int(num_match.group(1))

    # "qual" (singular) sem _top = top 1
    if result and "_top" not in result:
        if any(t in tokens for t in ["qual"]):
            result["_top"] = 1

    return result


def apply_filters(data: list, filters: dict) -> list:
    """Aplica filtros, ordenacao e limite aos dados ja retornados."""
    if not data or not filters:
        return data

    sort_key = filters.pop("_sort", None)
    top_n = filters.pop("_top", None)
    result = data

    # Filtros especiais (funcoes)
    fn_keys = [k for k in filters if k.startswith("_fn_")]
    for fn_key in fn_keys:
        field_spec = filters.pop(fn_key)
        fn_name = fn_key.replace("_fn_", "")
        if fn_name == "empty":
            result = [r for r in result if isinstance(r, dict) and not str(r.get(field_spec, "") or "").strip()]
        elif fn_name == "not_empty":
            result = [r for r in result if isinstance(r, dict) and str(r.get(field_spec, "") or "").strip()]
        elif fn_name in ("maior", "menor") and ":" in str(field_spec):
            campo, valor_str = str(field_spec).split(":", 1)
            try:
                threshold = float(valor_str)
                if fn_name == "maior":
                    result = [r for r in result if isinstance(r, dict) and float(r.get(campo, 0) or 0) > threshold]
                else:
                    result = [r for r in result if isinstance(r, dict) and float(r.get(campo, 0) or 0) < threshold]
            except (ValueError, TypeError):
                pass
        elif fn_name == "contem" and ":" in str(field_spec):
            campo, texto = str(field_spec).split(":", 1)
            result = [r for r in result if isinstance(r, dict) and texto.upper() in str(r.get(campo, "")).upper()]

    # Filtrar por campo=valor
    for field, value in filters.items():
        if field.startswith("_"):
            continue
        result = [r for r in result if isinstance(r, dict) and str(r.get(field, "")).upper() == value.upper()]

    # Ordenar
    if sort_key and result:
        field, direction = sort_key.rsplit("_", 1)
        reverse = direction == "DESC"
        try:
            # Tentar sort numerico primeiro
            result = sorted(result, key=lambda r: float(r.get(field, 0) or 0), reverse=reverse)
        except (ValueError, TypeError):
            try:
                # Fallback: sort por string (funciona pra datas dd/mm/yyyy invertendo pra yyyy-mm-dd)
                def _sort_key(r):
                    v = str(r.get(field, "") or "")
                    # Converter dd/mm/yyyy pra yyyy-mm-dd pra sort correto
                    if re.match(r'\d{2}/\d{2}/\d{4}', v):
                        parts = v.split("/")
                        return f"{parts[2]}-{parts[1]}-{parts[0]}"
                    return v
                result = sorted(result, key=_sort_key, reverse=reverse)
            except Exception:
                pass

    # Limitar quantidade (top N)
    if top_n and result:
        result = result[:top_n]

    return result


# ============================================================
# DAILY TRAINING (scheduler de madrugada)
# ============================================================

TRAINING_HOUR = int(os.getenv("TRAINING_HOUR", "3"))

async def daily_training(force: bool = False) -> dict:
    """Executa compilacao + review de aliases via pool_train.
    Chamado automaticamente pelo scheduler ou manualmente via CLI/endpoint."""
    from src.llm.knowledge_compiler import KnowledgeCompiler

    stats = {"compiler": {}, "aliases_reviewed": 0, "error": None}
    print(f"[TRAIN] Iniciando treinamento {'(forcado)' if force else '(scheduled)'} ...")

    # 1. Knowledge Compiler
    try:
        compiler = KnowledgeCompiler(groq_api_key=pool_train.get_key() if pool_train.available else None)
        result = await compiler.compile(full=force, dry_run=False, verbose=True)
        stats["compiler"] = result

        # Recarregar no SmartAgent
        if result.get("processed", 0) > 0:
            global _COMPILED_LOADED
            _COMPILED_LOADED = False
            _load_compiled_knowledge()
            print(f"[TRAIN] Knowledge recarregado ({result['processed']} docs)")
    except Exception as e:
        stats["compiler"] = {"error": str(e)}
        print(f"[TRAIN] Compiler falhou: {e}")

    # 2. Alias review (aprovar sugestoes de alta confianca)
    try:
        from src.llm.alias_resolver import AliasResolver
        ar = AliasResolver()
        suggestions = ar.get_suggestions("pending")
        auto_approved = 0
        for s in suggestions:
            if s.get("confidence", 0) >= 0.85 and s.get("count", 0) >= 3:
                ar.approve_suggestion(s["apelido"], nome_real=s.get("nome_real"), codprod=s.get("codprod"))
                auto_approved += 1
                print(f"[TRAIN] Auto-aprovado alias: {s['apelido']} -> {s.get('nome_real', s.get('codprod'))}")
        stats["aliases_reviewed"] = auto_approved
    except Exception as e:
        print(f"[TRAIN] Alias review falhou: {e}")

    print(f"[TRAIN] Concluido: compiler={stats['compiler'].get('processed', 0)} docs, aliases={stats['aliases_reviewed']}")
    return stats


async def _training_scheduler():
    """Loop infinito que roda daily_training() no horario configurado."""
    while True:
        now = datetime.now()
        # Calcular proximo horario
        target = now.replace(hour=TRAINING_HOUR, minute=0, second=0, microsecond=0)
        if now >= target:
            target = target.replace(day=target.day + 1)
        wait_seconds = (target - now).total_seconds()
        print(f"[TRAIN] Proximo treino em {wait_seconds/3600:.1f}h ({target.strftime('%d/%m %H:%M')})")
        await asyncio.sleep(wait_seconds)
        try:
            await daily_training()
        except Exception as e:
            print(f"[TRAIN] Erro no scheduler: {e}")


class ConversationContext:
    """Contexto de conversa de um usuario. Guarda parametros e dados anteriores."""

    def __init__(self, user_id: str):
        self.user_id = user_id
        self.intent = None          # ultimo intent resolvido
        self.params = {}            # {marca, fornecedor, empresa, comprador, periodo}
        self.last_result = {}       # ultimo resultado (detail_data, columns, etc)
        self.last_question = ""     # ultima pergunta
        self.last_view_mode = "pedidos"
        self.turn_count = 0         # quantas perguntas ja fez
        self._extra_columns = []    # colunas extras pedidas pelo usuario

    def merge_params(self, new_params: dict) -> dict:
        """Mescla parametros novos com contexto anterior.
        Regra: parametro novo sobrescreve, ausente herda do contexto.
        """
        merged = {}
        param_keys = ["marca", "fornecedor", "empresa", "comprador", "periodo",
                       "codprod", "produto_nome", "pedido"]

        for key in param_keys:
            new_val = new_params.get(key)
            old_val = self.params.get(key)

            if new_val:
                merged[key] = new_val  # Novo sobrescreve
            elif old_val:
                merged[key] = old_val  # Herda do contexto

        return merged

    def update(self, intent: str, params: dict, result: dict, question: str, view_mode: str = "pedidos"):
        """Atualiza contexto apos uma resposta bem-sucedida."""
        if intent != self.intent:
            self._extra_columns = []  # Limpa ao mudar de intent
        self.intent = intent
        # Atualiza params (nao apaga os antigos, so sobrescreve os que vieram)
        for k, v in params.items():
            if v:
                self.params[k] = v
        self.last_result = result
        self.last_question = question
        self.last_view_mode = view_mode
        self.turn_count += 1

    def has_data(self) -> bool:
        """Tem dados anteriores disponíveis para filtrar?"""
        return bool(self.last_result and self.last_result.get("detail_data"))

    def get_data(self) -> list:
        """Retorna dados anteriores."""
        return self.last_result.get("detail_data", [])

    def get_description(self) -> str:
        """Retorna descrição do ultimo resultado."""
        return self.last_result.get("description", "")

    def __repr__(self):
        return f"<Ctx user={self.user_id} intent={self.intent} params={self.params} turns={self.turn_count}>"


# ============================================================
# SMART AGENT v3
# ============================================================

class SmartAgent:
    def __init__(self):
        self.executor = SafeQueryExecutor()
        self._known_marcas = set()
        self._known_empresas = set()
        self._known_compradores = set()
        self._entities_loaded = False
        self.kb = KnowledgeBase()  # Knowledge Base para perguntas sobre processos
        self.alias_resolver = AliasResolver()  # Apelidos de produto
        self.query_logger = QueryLogger()
        self.result_validator = ResultValidator()
        # Contexto POR USUARIO
        self._user_contexts = {}  # {user_id: ConversationContext}
        # Carregar knowledge compilado (auto-gerado)
        _load_compiled_knowledge()

    def _get_context(self, user_id: str) -> 'ConversationContext':
        """Retorna (ou cria) contexto do usuario."""
        if not user_id:
            user_id = "__default__"
        if user_id not in self._user_contexts:
            self._user_contexts[user_id] = ConversationContext(user_id)
        return self._user_contexts[user_id]

    def clear_user(self, user_id: str):
        """Limpa contexto de um usuario especifico (logout)."""
        if user_id in self._user_contexts:
            del self._user_contexts[user_id]
            print(f"[CTX] Contexto limpo: {user_id}")

    def clear(self):
        """Limpa todos os contextos (compatibilidade)."""
        self._user_contexts.clear()
        print(f"[CTX] Todos os contextos limpos")

    async def _load_entities(self):
        """Carrega entidades do banco (1x, depois cache)."""
        if self._entities_loaded:
            return
        try:
            # Marcas
            r = await self.executor.execute("SELECT UPPER(TRIM(DESCRICAO)) AS M FROM TGFMAR WHERE DESCRICAO IS NOT NULL")
            if r.get("success"):
                for row in r.get("data", []):
                    v = row.get("M", row[0] if isinstance(row, (list, tuple)) else "") if isinstance(row, dict) else (str(row[0]) if isinstance(row, (list, tuple)) and row else "")
                    if v and len(v) > 1: self._known_marcas.add(v.strip())
            # Empresas
            r = await self.executor.execute("SELECT UPPER(TRIM(NOMEFANTASIA)) AS E FROM TSIEMP WHERE NOMEFANTASIA IS NOT NULL")
            if r.get("success"):
                for row in r.get("data", []):
                    v = row.get("E", row[0] if isinstance(row, (list, tuple)) else "") if isinstance(row, dict) else (str(row[0]) if isinstance(row, (list, tuple)) and row else "")
                    if v and len(v) > 1: self._known_empresas.add(v.strip())
            # Compradores (via marcas)
            r = await self.executor.execute("SELECT DISTINCT UPPER(TRIM(V.APELIDO)) AS C FROM TGFVEN V JOIN TGFMAR M ON M.AD_CODVEND = V.CODVEND WHERE V.APELIDO IS NOT NULL")
            if r.get("success"):
                for row in r.get("data", []):
                    v = row.get("C", row[0] if isinstance(row, (list, tuple)) else "") if isinstance(row, dict) else (str(row[0]) if isinstance(row, (list, tuple)) and row else "")
                    if v and len(v) > 1: self._known_compradores.add(v.strip())
            print(f"[SMART] Entidades: {len(self._known_marcas)} marcas, {len(self._known_empresas)} empresas, {len(self._known_compradores)} compradores")
        except Exception as e:
            print(f"[SMART] Erro ao carregar entidades: {e}")
        self._entities_loaded = True

    async def ask(self, question: str, user_context: dict = None) -> Optional[dict]:
        """Entry point com logging integrado."""
        user_id = (user_context or {}).get("user", "")
        _log = self.query_logger.create_entry(question, user=user_id)

        try:
            result = await self._ask_core(question, user_context, _log)
        except Exception as e:
            print(f"[SMART] Erro no ask: {e}")
            result = self._handle_fallback(question)
            _log["processing"]["layer"] = "error"

        # Finalizar log e salvar
        try:
            _detail_data = result.pop("_detail_data", None) if result else None
            self._finalize_log(_log, result, _detail_data)
            self.query_logger.save(_log)
            # Recolocar _detail_data para o endpoint montar table_data (toggle frontend)
            if _detail_data and result:
                result["_detail_data"] = _detail_data
        except Exception:
            pass

        if result:
            result["message_id"] = _log["id"]

            # B2: Registrar termos sem match como sugestao de alias
            query_results = result.get("query_results")
            produto_nome_used = getattr(self, '_last_produto_nome', None)

            if query_results == 0 and produto_nome_used:
                self.alias_resolver.suggest_alias(
                    produto_nome_used,
                    context=question,
                    user=user_id
                )
                # B4: Salvar failed term para deteccao de sequencia
                if not hasattr(self, '_failed_terms'):
                    self._failed_terms = {}
                self._failed_terms[user_id] = produto_nome_used
            elif query_results and query_results > 0 and produto_nome_used:
                # B4: Sequencia detectada - query anterior falhou, esta passou
                failed = getattr(self, '_failed_terms', {}).get(user_id)
                if failed and failed != produto_nome_used:
                    codprod_found = None
                    detail = result.get("_detail_data") or []
                    if detail and isinstance(detail[0], dict):
                        codprod_found = detail[0].get("CODPROD")
                    self.alias_resolver.detect_alias_from_sequence(failed, produto_nome_used, codprod_found)
                    self._failed_terms.pop(user_id, None)

            self._last_produto_nome = None

            # Prefixar resposta com alias resolvido
            if hasattr(self, '_last_alias_resolved') and self._last_alias_resolved:
                alias_info = self._last_alias_resolved
                self._last_alias_resolved = None
                resp = result.get("response", "")
                prefix = f"*'{alias_info['from']}' = {alias_info['to']}*\n\n"
                result["response"] = prefix + resp
        return result

    def _finalize_log(self, _log: dict, result: dict, _detail_data: list = None):
        """Preenche metadata final do log a partir do resultado."""
        if not result:
            return
        proc = _log.get("processing", {})
        proc["time_ms"] = result.get("time_ms", 0)
        # Result type mapping
        tipo = result.get("tipo", "info")
        type_map = {"consulta_banco": "table", "info": "knowledge", "arquivo": "table", "erro": "error"}
        records = result.get("query_results") or 0
        _log["result"] = {
            "type": type_map.get(tipo, tipo),
            "records_found": records,
            "records_shown": records,
            "response_preview": (result.get("response") or "")[:200],
        }
        _log["auto_tags"] = generate_auto_tags(_log)

        # Result data summary (para validacao offline)
        if _detail_data and isinstance(_detail_data, list):
            _log["result_data_summary"] = build_result_data_summary(_detail_data, _log)

        # Auto-auditoria: validacao em tempo real
        try:
            validation = self.result_validator.validate(_log, _detail_data)
            _log["validation"] = {
                "passed": validation["passed"],
                "checks_run": validation["checks_run"],
                "checks_failed": validation["checks_failed"],
                "severity": validation.get("severity"),
                "suggested_fix": validation.get("suggested_fix"),
                "checks": validation.get("checks", []),
            }
            # Console warnings para severity alta
            if validation.get("severity") in ("high", "critical"):
                for check in validation.get("checks", []):
                    if not check.get("passed"):
                        print(f"[AUDIT] {check.get('check', '?')}: {check.get('detail', '?')}")
        except Exception as e:
            print(f"[AUDIT] Erro na validacao: {e}")

    async def _ask_core(self, question: str, user_context: dict = None, _log: dict = None) -> Optional[dict]:
        await self._load_entities()
        t0 = time.time()
        tokens = tokenize(question)
        q_norm = normalize(question)
        user_id = (user_context or {}).get("user", "__default__")
        ctx = self._get_context(user_id)

        # Score de cada intent
        scores = score_intent(tokens)
        best_intent = max(scores, key=scores.get)
        best_score = scores[best_intent]

        if _log:
            _log["processing"]["score"] = best_score

        # ========== CONFIRMACAO CURTA (follow-up excel) ==========
        if len(tokens) <= 3 and any(t in CONFIRM_WORDS for t in tokens) and ctx.has_data():
            print(f"[SMART] Follow-up: gerar_excel")
            if _log: _log["processing"].update(layer="scoring", intent="gerar_excel")
            return await self._handle_excel_followup(user_context, ctx)

        # Excel explicito
        if scores.get("gerar_excel", 0) >= INTENT_THRESHOLDS["gerar_excel"]:
            if ctx.has_data():
                if _log: _log["processing"].update(layer="scoring", intent="gerar_excel")
                return await self._handle_excel_followup(user_context, ctx)

        # Saudacao (so se for curta e score alto)
        if scores.get("saudacao", 0) >= INTENT_THRESHOLDS["saudacao"] and len(tokens) <= 5:
            if _log: _log["processing"].update(layer="scoring", intent="saudacao")
            return self._handle_saudacao(user_context)

        # Ajuda
        if scores.get("ajuda", 0) >= INTENT_THRESHOLDS["ajuda"]:
            if _log: _log["processing"].update(layer="scoring", intent="ajuda")
            return self._handle_ajuda()

        # ========== DETECTAR FOLLOW-UP (referencia a dados anteriores) ==========
        is_followup = detect_followup(tokens, q_norm)
        filters = detect_filter_request(q_norm, tokens) if is_followup else {}

        if is_followup and ctx.has_data() and filters:
            # Filtrar dados anteriores
            print(f"[CTX] Follow-up com filtro: {filters} | ctx={ctx}")
            if _log: _log["processing"].update(layer="scoring", intent=ctx.intent or "follow_up", filters_source="pattern", filters_applied={k: str(v) for k, v in filters.items()})
            result = self._handle_filter_followup(ctx, filters, question, t0)
            if result:
                return result

        # ========== EXTRAIR PARAMETROS + MERGE COM CONTEXTO ==========
        params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        has_entity = params.get("marca") or params.get("fornecedor") or params.get("comprador") or params.get("empresa")

        # Se e follow-up sem entidade nova, herda do contexto
        if is_followup and not has_entity and ctx.params:
            merged = ctx.merge_params(params)
            if merged != params:
                print(f"[CTX] Herdando contexto: {params} -> {merged}")
                params = merged
                has_entity = params.get("marca") or params.get("fornecedor") or params.get("comprador") or params.get("empresa")

        # Se score baixo mas tem contexto e a pergunta parece continuacao
        if best_score < INTENT_THRESHOLDS.get(best_intent, 8) and is_followup and ctx.intent:
            # Herdar intent anterior se nenhum novo foi detectado forte
            if best_score <= 3:
                best_intent = ctx.intent
                best_score = INTENT_THRESHOLDS.get(best_intent, 8)  # Forca threshold
                print(f"[CTX] Herdando intent: {best_intent} (follow-up sem score)")

        # Salvar produto_nome original para B2 (registro de termos sem match)
        self._last_produto_nome = params.get("produto_nome")

        # ========== LAYER 0.4: ALIAS RESOLVER ==========
        self._last_alias_resolved = None
        if params.get("produto_nome"):
            alias = self.alias_resolver.resolve(params["produto_nome"])
            if alias:
                original_term = params["produto_nome"]
                resolved_to = alias.get("nome_real") or str(alias.get("codprod"))
                if alias.get("codprod"):
                    params["codprod"] = alias["codprod"]
                    params.pop("produto_nome", None)
                elif alias.get("nome_real"):
                    params["produto_nome"] = alias["nome_real"]
                self._last_alias_resolved = {"from": original_term, "to": resolved_to}
                print(f"[ALIAS] Resolvido: '{original_term}' -> {resolved_to} (conf={alias.get('confidence',0):.0%})")

        # ========== LAYER 0.5: PRODUTO (roteamento especial) ==========
        product_type = detect_product_query(q_norm, params)
        if product_type:
            print(f"[SMART] Layer 0.5 (produto): type={product_type} | params={params}")
            if _log: _log["processing"].update(layer="produto", intent=product_type)
            if product_type == "busca_fabricante":
                return await self._handle_busca_fabricante(question, user_context, t0, params, ctx)
            elif product_type == "similares":
                return await self._handle_similares(question, user_context, t0, params, ctx)
            elif product_type == "produto_360":
                return await self._handle_produto_360(question, user_context, t0, params, ctx)
            elif product_type == "busca_aplicacao":
                return await self._handle_busca_aplicacao(question, user_context, t0, params, ctx)

        # ========== LAYER 1: SCORING (0ms) ==========
        print(f"[SMART] Scores: pend={scores.get('pendencia_compras',0)} est={scores.get('estoque',0)} vend={scores.get('vendas',0)} | best={best_intent}({best_score}) | followup={is_followup}")

        # Log: entities extraidas
        if _log:
            _log["processing"]["entities"] = {k: v for k, v in params.items() if v and k != "periodo"}

        # Checar knowledge score antes pra decidir rota
        kb_score = score_knowledge(question)

        # Se KB score alto E data score nao e dominante, vai pra knowledge
        if kb_score >= 8 and (best_score < kb_score or best_score < 12):
            print(f"[SMART] Layer 1.2 (knowledge): kb={kb_score} vs data={best_score}")
            if _log: _log["processing"].update(layer="scoring", intent="conhecimento")
            kb_result = await self.kb.answer(question)
            if kb_result:
                kb_result["time_ms"] = int((time.time() - t0) * 1000)
                return kb_result

        if best_score >= INTENT_THRESHOLDS.get(best_intent, 8):
            # Score alto = confianca alta no INTENT, mas query pode ter filtros complexos
            print(f"[SMART] Layer 1 (scoring): {best_intent} (score={best_score})")
            if _log: _log["processing"].update(layer="scoring", intent=best_intent)

            # Detectar se query e complexa (filtros/ordenacao que o pattern nao pegou)
            pattern_filters = detect_filter_request(q_norm, tokens)

            # Intent override: se filter rules detectou TIPO_COMPRA, forcar pendencia_compras
            # (evita "compras de estoque da eaton" cair em intent=estoque)
            if pattern_filters.get("TIPO_COMPRA") and best_intent != "pendencia_compras":
                print(f"[SMART] Intent override: {best_intent} -> pendencia_compras (TIPO_COMPRA detectado)")
                best_intent = "pendencia_compras"
                if _log: _log["processing"]["intent"] = "pendencia_compras"
                # Propagar tipo_compra para params (usado no _build_where_extra)
                params["tipo_compra"] = pattern_filters["TIPO_COMPRA"]

            is_complex = _is_complex_query(q_norm, tokens, pattern_filters)

            if is_complex and best_intent in ("pendencia_compras", "estoque", "vendas") and pool_classify.available:
                # Query complexa: usar Groq pra interpretar filtros (scoring ja resolveu intent)
                print(f"[SMART] Layer 1+ (Groq filtro): query complexa, consultando Groq...")
                llm_result = await groq_classify(question)
                if llm_result:
                    if _log: _log["processing"].update(layer="scoring+groq", groq_raw=dict(llm_result))
                    llm_filters = _llm_to_filters(llm_result, question)
                    # Usar entidades da LLM se scoring nao extraiu direito
                    for key in ["marca", "fornecedor", "empresa", "comprador", "periodo", "aplicacao"]:
                        if llm_result.get(key) and not params.get(key):
                            params[key] = llm_result[key]
                        elif llm_result.get(key) and params.get(key):
                            llm_val = llm_result[key].upper()
                            cur_val = params[key].upper()
                            if key == "marca" and self._known_marcas:
                                if llm_val in self._known_marcas and cur_val not in self._known_marcas:
                                    print(f"[SMART] LLM corrigiu {key}: {cur_val} -> {llm_val}")
                                    params[key] = llm_val
                                    if _log: _log["processing"]["groq_corrected"] = True
                    # Extrair e normalizar extra_columns
                    extra_columns = []
                    raw_extra = llm_result.get("extra_columns") or []
                    for col in raw_extra:
                        norm = COLUMN_NORMALIZE.get(col.upper().replace(" ", "_"), col.upper())
                        if norm not in extra_columns:
                            extra_columns.append(norm)
                    if extra_columns:
                        print(f"[SMART] Extra columns: {extra_columns}")
                        if _log: _log["processing"]["extra_columns"] = extra_columns
                    view_mode = llm_result.get("view") or detect_view_mode(tokens)
                    if _log:
                        _log["processing"]["view_mode"] = view_mode
                        _log["processing"]["entities"] = {k: v for k, v in params.items() if v and k != "periodo"}
                    if llm_filters:
                        print(f"[SMART] LLM filters: {llm_filters}")
                        if _log: _log["processing"].update(filters_source="llm", filters_applied={k: str(v) for k, v in llm_filters.items()})
                    if best_intent == "pendencia_compras":
                        return await self._handle_pendencia_compras(question, user_context, t0, params, view_mode, ctx, llm_filters=llm_filters, extra_columns=extra_columns)
                    elif best_intent == "estoque":
                        return await self._handle_estoque(question, user_context, t0, params, ctx)
                    elif best_intent == "vendas":
                        return await self._handle_vendas(question, user_context, t0, params, ctx)

            if _log and pattern_filters:
                _log["processing"].update(filters_source="pattern", filters_applied={k: str(v) for k, v in pattern_filters.items()})
            # Propagar tipo_compra do filter para params (SQL WHERE)
            if pattern_filters.get("TIPO_COMPRA") and not params.get("tipo_compra"):
                params["tipo_compra"] = pattern_filters["TIPO_COMPRA"]
            return await self._dispatch(best_intent, question, user_context, t0, tokens, params, ctx)

        # ========== LAYER 1.5: ENTITY DETECTION ==========
        if has_entity and best_score >= 3:
            # Tem entidade + algum score = provavelmente pendencia
            print(f"[SMART] Layer 1.5 (entidade + score): pendencia | params={params}")
            if _log: _log["processing"].update(layer="scoring", intent="pendencia_compras")
            view_mode = detect_view_mode(tokens)
            return await self._handle_pendencia_compras(question, user_context, t0, params, view_mode, ctx)

        # ========== LAYER 2: LLM CLASSIFIER (Groq ~0.5s / Ollama ~10s) ==========
        if USE_LLM_CLASSIFIER:
            print(f"[SMART] Layer 2 (LLM): score ambiguo ({best_score}), consultando LLM...")
            llm_result = await llm_classify(question)

            if llm_result and llm_result.get("intent") not in (None, "desconhecido", ""):
                intent = llm_result["intent"]
                print(f"[SMART] LLM classificou: {intent} | filtro={llm_result.get('filtro')} | ordenar={llm_result.get('ordenar')} | top={llm_result.get('top')}")

                if _log:
                    _layer = "groq" if pool_classify.available else "ollama"
                    _log["processing"].update(layer=_layer, intent=intent, groq_raw=dict(llm_result) if pool_classify.available else None)

                # Se LLM classificou como conhecimento
                if intent == "conhecimento":
                    kb_result = await self.kb.answer(question)
                    if kb_result:
                        kb_result["time_ms"] = int((time.time() - t0) * 1000)
                        return kb_result

                # Usar entidades da LLM se nao extraiu pelo scoring
                for key in ["marca", "fornecedor", "empresa", "comprador", "periodo", "aplicacao"]:
                    if llm_result.get(key) and not params.get(key):
                        params[key] = llm_result[key]

                # Extrair e normalizar extra_columns
                extra_columns = []
                raw_extra = llm_result.get("extra_columns") or []
                for col in raw_extra:
                    norm = COLUMN_NORMALIZE.get(col.upper().replace(" ", "_"), col.upper())
                    if norm not in extra_columns:
                        extra_columns.append(norm)
                if extra_columns:
                    print(f"[SMART] Extra columns (L2): {extra_columns}")
                    if _log: _log["processing"]["extra_columns"] = extra_columns

                view_mode = llm_result.get("view") or detect_view_mode(tokens)

                # Converter filtro/ordenar/top da LLM em dict de filtros
                llm_filters = _llm_to_filters(llm_result, question)
                if llm_filters:
                    print(f"[SMART] LLM filters convertidos: {llm_filters}")
                    if _log: _log["processing"].update(filters_source="llm", filters_applied={k: str(v) for k, v in llm_filters.items()})

                # Propagar tipo_compra da LLM para params (SQL WHERE)
                if llm_result.get("tipo_compra") and not params.get("tipo_compra"):
                    tc = llm_result["tipo_compra"].lower()
                    if tc in ("casada", "empenho", "vinculada"):
                        params["tipo_compra"] = "Casada"
                    elif tc in ("estoque", "futura", "reposicao"):
                        params["tipo_compra"] = "Estoque"
                if llm_filters.get("TIPO_COMPRA") and not params.get("tipo_compra"):
                    params["tipo_compra"] = llm_filters["TIPO_COMPRA"]
                if params.get("tipo_compra") and intent != "pendencia_compras":
                    print(f"[SMART] Intent override (LLM): {intent} -> pendencia_compras (tipo_compra={params['tipo_compra']})")
                    intent = "pendencia_compras"
                    if _log: _log["processing"]["intent"] = "pendencia_compras"

                if _log:
                    _log["processing"]["view_mode"] = view_mode
                    _log["processing"]["entities"] = {k: v for k, v in params.items() if v and k != "periodo"}

                if intent == "pendencia_compras":
                    return await self._handle_pendencia_compras(question, user_context, t0, params, view_mode, ctx, llm_filters=llm_filters, extra_columns=extra_columns)
                elif intent == "estoque":
                    return await self._handle_estoque(question, user_context, t0, params, ctx)
                elif intent == "vendas":
                    return await self._handle_vendas(question, user_context, t0, params, ctx)
                elif intent == "produto":
                    # Redirecionar para detect_product_query ou busca_fabricante como fallback
                    product_type = detect_product_query(q_norm, params)
                    if product_type == "similares":
                        return await self._handle_similares(question, user_context, t0, params, ctx)
                    elif product_type == "produto_360":
                        return await self._handle_produto_360(question, user_context, t0, params, ctx)
                    elif product_type == "busca_aplicacao":
                        return await self._handle_busca_aplicacao(question, user_context, t0, params, ctx)
                    else:
                        return await self._handle_busca_fabricante(question, user_context, t0, params, ctx)
                elif intent == "saudacao":
                    return self._handle_saudacao(user_context)
                elif intent == "ajuda":
                    return self._handle_ajuda()

        # ========== LAYER 3: FALLBACK ==========
        # Ultima tentativa: se tem entidade, assume pendencia
        if has_entity:
            print(f"[SMART] Layer 3 (fallback c/ entidade): {params}")
            if _log: _log["processing"].update(layer="fallback", intent="pendencia_compras")
            view_mode = detect_view_mode(tokens)
            return await self._handle_pendencia_compras(question, user_context, t0, params, view_mode, ctx)

        # Ultima tentativa: se tem algum score de knowledge, tenta KB
        if kb_score >= 4:
            print(f"[SMART] Layer 3 (fallback knowledge): kb_score={kb_score}")
            if _log: _log["processing"].update(layer="fallback", intent="conhecimento")
            kb_result = await self.kb.answer(question)
            if kb_result:
                kb_result["time_ms"] = int((time.time() - t0) * 1000)
                return kb_result

        if _log: _log["processing"]["layer"] = "fallback"
        return self._handle_fallback(question)

    async def _dispatch(self, intent: str, question: str, user_context: dict, t0: float, tokens: list, params: dict = None, ctx: ConversationContext = None):
        """Despacha para o handler correto baseado no intent."""
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        if intent == "pendencia_compras":
            view_mode = detect_view_mode(tokens)
            return await self._handle_pendencia_compras(question, user_context, t0, params, view_mode, ctx)
        elif intent == "estoque":
            return await self._handle_estoque(question, user_context, t0, params, ctx)
        elif intent == "vendas":
            return await self._handle_vendas(question, user_context, t0, params, ctx)
        elif intent == "produto":
            q_norm = normalize(question)
            product_type = detect_product_query(q_norm, params)
            if product_type == "similares":
                return await self._handle_similares(question, user_context, t0, params, ctx)
            elif product_type == "produto_360":
                return await self._handle_produto_360(question, user_context, t0, params, ctx)
            elif product_type == "busca_aplicacao":
                return await self._handle_busca_aplicacao(question, user_context, t0, params, ctx)
            else:
                return await self._handle_busca_fabricante(question, user_context, t0, params, ctx)
        elif intent == "saudacao":
            return self._handle_saudacao(user_context)
        elif intent == "ajuda":
            return self._handle_ajuda()
        return self._handle_fallback(question)

    # ---- FILTRO FOLLOW-UP (dados anteriores) ----
    def _handle_filter_followup(self, ctx: ConversationContext, filters: dict, question: str, t0: float):
        """Filtra dados anteriores baseado na pergunta do usuario."""
        data = ctx.get_data()
        if not data:
            return None

        # Guardar meta antes de apply_filters (que faz pop)
        has_sort = "_sort" in filters
        has_top = "_top" in filters
        top_n = filters.get("_top", 0)
        sort_key = filters.get("_sort", "")
        filter_fields = {k: v for k, v in filters.items() if not k.startswith("_")}
        fn_fields = {k: v for k, v in filters.items() if k.startswith("_fn_")}

        filtered = apply_filters(list(data), dict(filters))

        if not filtered:
            desc = ctx.get_description()
            return {
                "response": f"\U0001f914 Nenhum resultado encontrado com esse filtro nos dados de **{desc}**.\n\nTente outra pergunta ou faca uma nova consulta.",
                "tipo": "consulta_banco",
                "query_executed": f"Filtro: {filters}",
                "query_results": 0,
                "time_ms": int((time.time() - t0) * 1000),
            }

        desc = ctx.get_description()
        prev_intent = ctx.intent or "pendencia_compras"

        # Construir descricao do filtro para o titulo
        filter_parts = []
        if filter_fields:
            filter_parts.append(", ".join(f"{v}" for v in filter_fields.values()))
        for fn_key, fn_field in fn_fields.items():
            fn_name = fn_key.replace("_fn_", "")
            if fn_name == "empty":
                filter_parts.append(f"sem {fn_field.lower().replace('_', ' ')}")
        if has_sort and sort_key:
            sort_names = {
                "DIAS_ABERTO_DESC": "mais atrasado", "DIAS_ABERTO_ASC": "mais recente",
                "VLR_PENDENTE_DESC": "maior valor", "VLR_PENDENTE_ASC": "menor valor",
                "DT_PEDIDO_DESC": "mais recente", "QTD_PENDENTE_DESC": "maior quantidade",
            }
            filter_parts.append(sort_names.get(sort_key, sort_key))

        filter_label = " | ".join(filter_parts) if filter_parts else "filtrado"

        if prev_intent == "pendencia_compras":
            # Resposta especial para top 1 (resposta direta)
            if top_n == 1 and len(filtered) == 1:
                r = filtered[0]
                pedido = r.get("PEDIDO", "?")
                produto = r.get("PRODUTO", "?")
                marca = r.get("MARCA", "")
                vlr = float(r.get("VLR_PENDENTE", 0) or 0)
                dias = int(r.get("DIAS_ABERTO", 0) or 0)
                qtd = int(r.get("QTD_PENDENTE", 0) or 0)
                status = r.get("STATUS_ENTREGA", "?")
                fornecedor = r.get("FORNECEDOR", "")

                response = f"\U0001f50d **{filter_label.title()}** de **{desc.title()}**:\n\n"
                response += f"| Campo | Valor |\n|---|---|\n"
                response += f"| **Pedido** | {pedido} |\n"
                if fornecedor: response += f"| **Fornecedor** | {fornecedor} |\n"
                response += f"| **Produto** | {produto} |\n"
                if marca: response += f"| **Marca** | {marca} |\n"
                response += f"| **Qtd. Pendente** | {fmt_num(qtd)} |\n"
                response += f"| **Valor Pendente** | R$ {fmt_num(vlr)} |\n"
                response += f"| **Dias em Aberto** | {dias} dias |\n"
                response += f"| **Status** | {status} |\n"
            else:
                # Tabela normal com KPIs
                view_mode = "itens"
                kpis_data = [{
                    "QTD_PEDIDOS": len(set(str(r.get("PEDIDO", "")) for r in filtered if isinstance(r, dict))),
                    "QTD_ITENS": len(filtered),
                    "VLR_PENDENTE": sum(float(r.get("VLR_PENDENTE", 0) or 0) for r in filtered if isinstance(r, dict)),
                }]
                response = format_pendencia_response(kpis_data, filtered, f"{desc} ({filter_label})", ctx.params, view_mode, extra_columns=ctx._extra_columns if ctx else [])
        elif prev_intent == "estoque":
            response = format_estoque_response(filtered, ctx.params)
        elif prev_intent == "vendas":
            kpi_row = {
                "QTD_VENDAS": len(filtered),
                "FATURAMENTO": sum(float(r.get("FATURAMENTO", 0) or 0) for r in filtered if isinstance(r, dict)),
                "TICKET_MEDIO": 0,
            }
            if kpi_row["QTD_VENDAS"]:
                kpi_row["TICKET_MEDIO"] = kpi_row["FATURAMENTO"] / kpi_row["QTD_VENDAS"]
            response = format_vendas_response(kpi_row, f"{desc} ({filter_label})")
        else:
            response = f"Encontrei **{len(filtered)}** registros com o filtro aplicado."

        elapsed = int((time.time() - t0) * 1000)
        return {
            "response": response,
            "tipo": "consulta_banco",
            "query_executed": f"Filtro: {filter_label}",
            "query_results": len(filtered),
            "time_ms": elapsed,
        }

    # ---- SAUDACAO ----
    def _handle_saudacao(self, user_context=None):
        nome = (user_context or {}).get("user", "")
        hora = datetime.now().hour
        saud = "Bom dia" if hora < 12 else ("Boa tarde" if hora < 18 else "Boa noite")
        r = f"{saud}{', ' + nome if nome else ''}! \U0001f44b\n\n"
        r += "Como posso ajudar? Posso consultar:\n\n"
        r += "\U0001f4e6 **Pendencia de compras** - *\"pendencia da marca Donaldson\"*\n"
        r += "\U0001f4ca **Vendas e faturamento** - *\"vendas de hoje\"*\n"
        r += "\U0001f4cb **Estoque** - *\"estoque do produto 133346\"*\n\nE so perguntar!"
        return {"response": r, "tipo": "info", "query_executed": None, "query_results": None}

    def _handle_ajuda(self):
        r = "\U0001f916 **O que posso fazer:**\n\n"
        r += "**\U0001f4e6 Pendencia de Compras**\n- *\"Pendencia da marca Donaldson\"*\n- *\"Itens pendentes da Tome\"*\n- *\"Pedidos em aberto do comprador Juliano\"*\n\n"
        r += "**\U0001f4ca Vendas**\n- *\"Vendas de hoje\"* / *\"Faturamento do mes\"*\n\n"
        r += "**\U0001f4cb Estoque**\n- *\"Estoque do produto 133346\"*\n- *\"Estoque critico\"*\n\n"
        r += "**\U0001f4da Processos e Regras**\n- *\"Como funciona o fluxo de compras?\"*\n- *\"O que e compra casada?\"*\n- *\"Qual a diferenca entre TOP 1301 e 1313?\"*\n\n"
        r += "**\U0001f4e5 Excel**\n- Apos qualquer consulta, diga *\"sim\"* ou *\"gera excel\"*\n\n"
        r += "\U0001f4a1 Tambem temos **Relatorios** completos no menu lateral!"
        return {"response": r, "tipo": "info", "query_executed": None, "query_results": None}

    def _handle_fallback(self, question):
        r = "\U0001f914 Nao entendi a pergunta.\n\nTente algo como:\n- *\"Pendencia da marca Donaldson\"*\n- *\"Vendas de hoje\"*\n- *\"Estoque do produto 133346\"*\n\nOu digite **ajuda** para ver tudo que posso fazer."
        return {"response": r, "tipo": "info", "query_executed": None, "query_results": None}

    # ---- PENDENCIA ----
    async def _handle_pendencia_compras(self, question, user_context, t0, params=None, view_mode="pedidos", ctx=None, llm_filters=None, extra_columns=None):
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        # Merge com contexto se disponivel
        if ctx and not (params.get("marca") or params.get("fornecedor") or params.get("comprador")):
            params = ctx.merge_params(params)
            print(f"[CTX] Params merged: {params}")

        # Herdar extra_columns do contexto se nao veio nesta pergunta
        if not extra_columns and ctx and hasattr(ctx, '_extra_columns') and ctx._extra_columns:
            extra_columns = ctx._extra_columns
            print(f"[CTX] Extra columns herdadas: {extra_columns}")

        extra_cols_normalized = extra_columns or []
        print(f"[SMART] Pendencia params: {params} | view: {view_mode} | llm_filters: {llm_filters} | extra_cols: {extra_cols_normalized}")

        sql_kpis, sql_detail, description = sql_pendencia_compras(params, user_context)

        # Verificar se precisa adicionar campos ao SQL
        needs_sql_extra = [c for c in extra_cols_normalized if c in EXTRA_SQL_FIELDS]
        if needs_sql_extra:
            extra_select = ", ".join(EXTRA_SQL_FIELDS[c] for c in needs_sql_extra)
            sql_detail = sql_detail.replace(
                f"\n        {JOINS_PENDENCIA}",
                f",\n            {extra_select}\n        {JOINS_PENDENCIA}"
            )
            print(f"[SMART] SQL com campos extras: {needs_sql_extra}")

        kpis_result = await self.executor.execute(sql_kpis)
        if not kpis_result.get("success"):
            return {"response": f"Erro ao consultar: {kpis_result.get('error','?')}", "tipo": "consulta_banco", "query_executed": sql_kpis[:200], "query_results": 0}

        kpis_data = kpis_result.get("data", [])
        if kpis_data and isinstance(kpis_data[0], (list, tuple)):
            cols = kpis_result.get("columns") or ["QTD_PEDIDOS", "QTD_ITENS", "VLR_PENDENTE"]
            kpis_data = [dict(zip(cols, row)) for row in kpis_data]

        qtd = int((kpis_data[0] if kpis_data else {}).get("QTD_PEDIDOS", 0) or 0)
        detail_data = []
        detail_columns = ["EMPRESA","PEDIDO","TIPO_COMPRA","COMPRADOR","DT_PEDIDO","PREVISAO_ENTREGA","CONFIRMADO","FORNECEDOR","CODPROD","PRODUTO","MARCA","APLICACAO","NUM_FABRICANTE","UNIDADE","QTD_PEDIDA","QTD_ATENDIDA","QTD_PENDENTE","VLR_UNITARIO","VLR_PENDENTE","DIAS_ABERTO","STATUS_ENTREGA"] + needs_sql_extra

        if qtd > 0:
            dr = await self.executor.execute(sql_detail)
            if dr.get("success"):
                detail_data = dr.get("data", [])
                if detail_data and isinstance(detail_data[0], (list, tuple)):
                    rc = dr.get("columns") or detail_columns
                    if rc and len(rc) == len(detail_data[0]):
                        detail_data = [dict(zip(rc, row)) for row in detail_data]
                    else:
                        detail_data = [dict(zip(detail_columns, row)) for row in detail_data]

        result_data = {"detail_data": detail_data, "columns": detail_columns, "description": description, "params": params, "intent": "pendencia_compras"}
        if ctx:
            ctx.update("pendencia_compras", params, result_data, question, view_mode)
            if extra_cols_normalized:
                ctx._extra_columns = extra_cols_normalized
            print(f"[CTX] Atualizado: {ctx}")

        # ========== VIEWS AGREGADAS (quem compra/fornece marca X?) ==========
        agg_view = detect_aggregation_view(normalize(question))
        if agg_view and detail_data:
            marca = params.get("marca", "")
            elapsed = int((time.time() - t0) * 1000)
            if agg_view == "comprador_marca" and marca:
                response = format_comprador_marca(detail_data, marca)
                return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": len(detail_data), "time_ms": elapsed, "_detail_data": detail_data}
            elif agg_view == "fornecedor_marca" and marca:
                response = format_fornecedor_marca(detail_data, marca)
                return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": len(detail_data), "time_ms": elapsed, "_detail_data": detail_data}

        # ========== FILTROS (LLM ou pattern-based) ==========
        # Prioridade: LLM filters > inline pattern filters
        active_filters = llm_filters if llm_filters else detect_filter_request(normalize(question), tokenize(question))
        filter_source = "LLM" if llm_filters else "pattern"

        if active_filters and detail_data:
            # Guardar meta antes de apply_filters
            has_top = "_top" in active_filters
            top_n = active_filters.get("_top", 0)
            sort_key = active_filters.get("_sort", "")
            filter_fields = {k: v for k, v in active_filters.items() if not k.startswith("_")}
            fn_fields = {k: v for k, v in active_filters.items() if k.startswith("_fn_")}

            filtered = apply_filters(list(detail_data), dict(active_filters))
            if filtered:
                print(f"[SMART] Filtro ({filter_source}): {active_filters} -> {len(filtered)}/{len(detail_data)} registros")

                # Construir label do filtro
                filter_parts = []
                if filter_fields:
                    filter_parts.append(", ".join(str(v) for v in filter_fields.values()))
                for fn_key, fn_field in fn_fields.items():
                    fn_name = fn_key.replace("_fn_", "")
                    if fn_name == "empty":
                        filter_parts.append(f"sem {fn_field.lower().replace('_', ' ')}")
                    elif fn_name == "not_empty":
                        filter_parts.append(f"com {fn_field.lower().replace('_', ' ')}")
                    elif fn_name in ("maior", "menor") and ":" in str(fn_field):
                        campo, val = str(fn_field).split(":", 1)
                        op = "acima de" if fn_name == "maior" else "abaixo de"
                        filter_parts.append(f"{campo.lower().replace('_', ' ')} {op} {val}")
                if sort_key:
                    sort_names = {"DIAS_ABERTO_DESC": "mais atrasado", "VLR_PENDENTE_DESC": "maior valor",
                                  "VLR_PENDENTE_ASC": "menor valor", "DT_PEDIDO_DESC": "pedido mais recente",
                                  "DT_PEDIDO_ASC": "pedido mais antigo",
                                  "PREVISAO_ENTREGA_DESC": "maior previsao de entrega",
                                  "PREVISAO_ENTREGA_ASC": "menor previsao de entrega",
                                  "QTD_PENDENTE_DESC": "maior quantidade"}
                    filter_parts.append(sort_names.get(sort_key, sort_key.lower().replace("_", " ")))
                filter_label = " | ".join(p for p in filter_parts if p)

                # Se view=pedidos e tem sort/top, agrupar por PEDIDO antes de aplicar top
                if view_mode == "pedidos" and top_n and sort_key:
                    # Pegar o pedido do primeiro item (ja esta ordenado)
                    top_pedido = str(filtered[0].get("PEDIDO", ""))
                    # Buscar TODOS os itens desse pedido nos dados originais
                    pedido_items = [r for r in detail_data if isinstance(r, dict) and str(r.get("PEDIDO", "")) == top_pedido]
                    if pedido_items:
                        filtered = pedido_items
                        print(f"[SMART] Pedido view: pedido {top_pedido} tem {len(pedido_items)} itens")

                # Top 1 + view pedidos = mostrar o PEDIDO completo
                if top_n == 1 and view_mode == "pedidos" and len(filtered) >= 1:
                    pedido_num = str(filtered[0].get("PEDIDO", "?"))
                    fornecedor = filtered[0].get("FORNECEDOR", "?")
                    dt_pedido = filtered[0].get("DT_PEDIDO", "?")
                    previsao = filtered[0].get("PREVISAO_ENTREGA", "?")
                    confirmado = filtered[0].get("CONFIRMADO", "?")
                    status = filtered[0].get("STATUS_ENTREGA", "?")
                    dias = int(filtered[0].get("DIAS_ABERTO", 0) or 0)

                    total_vlr = sum(float(r.get("VLR_PENDENTE", 0) or 0) for r in filtered)
                    total_itens = len(filtered)
                    total_qtd = sum(int(r.get("QTD_PENDENTE", 0) or 0) for r in filtered)

                    response = f"\U0001f50d **{filter_label.title() if filter_label else 'Resultado'}** de **{description.title()}**:\n\n"
                    response += f"| Campo | Valor |\n|---|---|\n"
                    response += f"| **Pedido** | {pedido_num} |\n"
                    response += f"| **Fornecedor** | {fornecedor} |\n"
                    response += f"| **Data Pedido** | {dt_pedido} |\n"
                    response += f"| **Previsao Entrega** | {previsao} |\n"
                    response += f"| **Confirmado** | {confirmado} |\n"
                    response += f"| **Status** | {status} |\n"
                    response += f"| **Dias em Aberto** | {dias} dias |\n"
                    response += f"| **Itens** | {total_itens} |\n"
                    response += f"| **Qtd. Pendente** | {fmt_num(total_qtd)} |\n"
                    response += f"| **Valor Pendente** | R$ {fmt_num(total_vlr)} |\n\n"

                    # Listar itens do pedido
                    if total_itens > 1:
                        response += f"**Itens do pedido {pedido_num}:**\n\n"
                        response += "| Produto | Marca | Qtd | Valor |\n|---|---|---|---|\n"
                        for r in filtered:
                            prod = str(r.get("PRODUTO", "?"))[:40]
                            marca_i = r.get("MARCA", "")
                            qtd_i = int(r.get("QTD_PENDENTE", 0) or 0)
                            vlr_i = float(r.get("VLR_PENDENTE", 0) or 0)
                            response += f"| {prod} | {marca_i} | {fmt_num(qtd_i)} | R$ {fmt_num(vlr_i)} |\n"

                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": total_itens, "time_ms": elapsed, "_detail_data": filtered}

                # Top 1 + view itens = mostrar 1 item (tabela vertical)
                elif top_n == 1 and len(filtered) == 1:
                    r = filtered[0]
                    response = f"\U0001f50d **{filter_label.title() if filter_label else 'Resultado'}** de **{description.title()}**:\n\n"
                    response += f"| Campo | Valor |\n|---|---|\n"
                    response += f"| **Pedido** | {r.get('PEDIDO', '?')} |\n"
                    if r.get("FORNECEDOR"): response += f"| **Fornecedor** | {r.get('FORNECEDOR')} |\n"
                    response += f"| **Produto** | {r.get('PRODUTO', '?')} |\n"
                    if r.get("MARCA"): response += f"| **Marca** | {r.get('MARCA')} |\n"
                    response += f"| **Qtd. Pendente** | {fmt_num(int(r.get('QTD_PENDENTE', 0) or 0))} |\n"
                    response += f"| **Valor Pendente** | R$ {fmt_num(float(r.get('VLR_PENDENTE', 0) or 0))} |\n"
                    response += f"| **Dias em Aberto** | {int(r.get('DIAS_ABERTO', 0) or 0)} dias |\n"
                    response += f"| **Status** | {r.get('STATUS_ENTREGA', '?')} |\n"
                    if r.get("PREVISAO_ENTREGA"): response += f"| **Previsao Entrega** | {r.get('PREVISAO_ENTREGA')} |\n"
                    if r.get("CONFIRMADO"): response += f"| **Confirmado** | {r.get('CONFIRMADO')} |\n"
                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": 1, "time_ms": elapsed, "_detail_data": filtered}
                else:
                    # Recalcular KPIs com dados filtrados
                    kpis_data = [{
                        "QTD_PEDIDOS": len(set(str(r.get("PEDIDO", "")) for r in filtered if isinstance(r, dict))),
                        "QTD_ITENS": len(filtered),
                        "VLR_PENDENTE": sum(float(r.get("VLR_PENDENTE", 0) or 0) for r in filtered if isinstance(r, dict)),
                    }]
                    detail_data = filtered
                    desc_filtered = f"{description} ({filter_label})" if filter_label else description
                    fallback_response = format_pendencia_response(kpis_data, detail_data, desc_filtered, params, "itens", extra_columns=extra_cols_normalized)
                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": fallback_response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": len(filtered), "time_ms": elapsed, "_detail_data": filtered}
            else:
                # Filtro retornou 0 resultados - avisar usuario
                print(f"[SMART] Filtro ({filter_source}): 0 resultados")
                filter_parts = []
                for k, v in filter_fields.items():
                    filter_parts.append(str(v))
                for fn_key, fn_field in fn_fields.items():
                    fn_name = fn_key.replace("_fn_", "")
                    if fn_name == "empty":
                        filter_parts.append(f"{fn_field} vazio")
                    elif fn_name in ("maior", "menor") and ":" in str(fn_field):
                        campo, val = str(fn_field).split(":", 1)
                        op = "acima de" if fn_name == "maior" else "abaixo de"
                        filter_parts.append(f"{campo} {op} {val}")
                filter_label = ", ".join(filter_parts) if filter_parts else "filtro aplicado"
                response = f"\U0001f4cb **{description.title()}**\n\n"
                response += f"\U0001f914 Nenhum item encontrado com **{filter_label}**.\n\n"
                response += f"Os {len(detail_data)} itens disponiveis tem os seguintes status: "
                statuses = set(str(r.get("STATUS_ENTREGA", "?")) for r in detail_data if isinstance(r, dict))
                response += ", ".join(f"**{s}**" for s in sorted(statuses)) + "."
                elapsed = int((time.time() - t0) * 1000)
                return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": 0, "time_ms": elapsed, "_detail_data": detail_data}

        fallback_response = format_pendencia_response(kpis_data, detail_data, description, params, view_mode, extra_columns=extra_cols_normalized)

        # Narrator: LLM explica os dados naturalmente
        if USE_LLM_NARRATOR and qtd > 0:
            summary = build_pendencia_summary(kpis_data, detail_data, params)
            narration = await llm_narrate(question, summary, "")
            if narration:
                # Monta resposta: analise da LLM + tabela resumida + oferta Excel
                parts = [f"\U0001f4e6 **{description.title()}**\n"]
                parts.append(narration)
                # Adicionar tabela resumida (sem KPIs, a LLM ja falou)
                table_lines = fallback_response.split("\n")
                table_start = next((i for i, l in enumerate(table_lines) if l.startswith("|")), None)
                if table_start is not None:
                    table_section = "\n".join(table_lines[table_start:])
                    parts.append(f"\n{table_section}")
                parts.append(f"\n\U0001f4e5 **Quer que eu gere um arquivo Excel com todos os {fmt_num(int((kpis_data[0] if kpis_data else {}).get('QTD_ITENS', 0) or 0))} itens?**")
                response = "\n".join(parts)
            else:
                response = fallback_response
        else:
            response = fallback_response

        elapsed = int((time.time() - t0) * 1000)
        return {
            "response": response,
            "tipo": "consulta_banco",
            "query_executed": sql_kpis[:200] + "...",
            "query_results": qtd,
            "time_ms": elapsed,
            "_detail_data": detail_data,
            "_visible_columns": extra_cols_normalized,
        }

    # ---- ESTOQUE ----
    async def _handle_estoque(self, question, user_context, t0, params=None, ctx=None):
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        # Merge com contexto se disponivel
        if ctx and not (params.get("codprod") or params.get("produto_nome") or params.get("marca")):
            params = ctx.merge_params(params)
        print(f"[SMART] Estoque params: {params}")
        q_norm = normalize(question)
        is_critico = any(w in q_norm for w in ["critico", "baixo", "zerado", "minimo", "acabando", "faltando"])

        if params.get("codprod"):
            sql = f"SELECT EMP.NOMEFANTASIA AS EMPRESA, E.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA, NVL(PRO.CARACTERISTICAS,'') AS APLICACAO, NVL(E.ESTOQUE,0) AS ESTOQUE, NVL(E.ESTMIN,0) AS ESTMIN FROM TGFEST E JOIN TGFPRO PRO ON PRO.CODPROD=E.CODPROD LEFT JOIN TGFMAR MAR ON MAR.CODIGO=PRO.CODMARCA LEFT JOIN TSIEMP EMP ON EMP.CODEMP=E.CODEMP WHERE E.CODPROD={params['codprod']} AND E.CODLOCAL=0 AND NVL(E.ATIVO,'S')='S' ORDER BY EMP.NOMEFANTASIA"
            dcols = ["EMPRESA","CODPROD","PRODUTO","MARCA","APLICACAO","ESTOQUE","ESTMIN"]
        elif params.get("produto_nome"):
            sql = f"SELECT EMP.NOMEFANTASIA AS EMPRESA, E.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA, NVL(PRO.CARACTERISTICAS,'') AS APLICACAO, NVL(E.ESTOQUE,0) AS ESTOQUE, NVL(E.ESTMIN,0) AS ESTMIN FROM TGFEST E JOIN TGFPRO PRO ON PRO.CODPROD=E.CODPROD LEFT JOIN TGFMAR MAR ON MAR.CODIGO=PRO.CODMARCA LEFT JOIN TSIEMP EMP ON EMP.CODEMP=E.CODEMP WHERE UPPER(PRO.DESCRPROD) LIKE UPPER('%{params['produto_nome']}%') AND E.CODLOCAL=0 AND NVL(E.ATIVO,'S')='S' AND NVL(E.ESTOQUE,0)>0 ORDER BY E.ESTOQUE DESC"
            dcols = ["EMPRESA","CODPROD","PRODUTO","MARCA","APLICACAO","ESTOQUE","ESTMIN"]
        elif is_critico or params.get("marca"):
            we = f"AND UPPER(MAR.DESCRICAO) LIKE UPPER('%{params['marca']}%')" if params.get("marca") else ""
            sql = f"SELECT E.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA, SUM(NVL(E.ESTOQUE,0)) AS ESTOQUE, MAX(NVL(E.ESTMIN,0)) AS ESTMIN FROM TGFEST E JOIN TGFPRO PRO ON PRO.CODPROD=E.CODPROD LEFT JOIN TGFMAR MAR ON MAR.CODIGO=PRO.CODMARCA WHERE E.CODLOCAL=0 AND NVL(E.ATIVO,'S')='S' AND NVL(E.ESTMIN,0)>0 {we} GROUP BY E.CODPROD, PRO.DESCRPROD, MAR.DESCRICAO HAVING SUM(NVL(E.ESTOQUE,0))<=MAX(NVL(E.ESTMIN,0)) ORDER BY SUM(NVL(E.ESTOQUE,0))"
            dcols = ["CODPROD","PRODUTO","MARCA","ESTOQUE","ESTMIN"]
        else:
            return {"response": "\U0001f4cb Para consultar estoque, me diga:\n\n- O **codigo** do produto: *\"estoque do produto 133346\"*\n- O **nome** da peca: *\"estoque da correia ventilador\"*\n- **Estoque critico**: *\"produtos com estoque critico\"*", "tipo": "info", "query_executed": None, "query_results": None}

        result = await self.executor.execute(sql)
        if not result.get("success"):
            return {"response": f"Erro ao consultar estoque: {result.get('error','?')}", "tipo": "consulta_banco", "query_executed": sql[:200], "query_results": 0}
        data = result.get("data", [])
        if data and isinstance(data[0], (list, tuple)):
            rc = result.get("columns") or dcols
            data = [dict(zip(rc if rc and len(rc)==len(data[0]) else dcols, row)) for row in data]

        result_data = {"detail_data": data, "columns": dcols, "description": "estoque", "params": params, "intent": "estoque"}
        if ctx:
            ctx.update("estoque", params, result_data, question)
        fallback_response = format_estoque_response(data, params)

        if USE_LLM_NARRATOR and data:
            summary = build_estoque_summary(data, params)
            narration = await llm_narrate(question, summary, "")
            if narration:
                table_lines = fallback_response.split("\n")
                table_start = next((i for i, l in enumerate(table_lines) if l.startswith("|")), None)
                parts = [narration]
                if table_start is not None:
                    parts.append(f"\n{''.join(table_lines[table_start:])}")
                response = "\n".join(parts)
            else:
                response = fallback_response
        else:
            response = fallback_response

        elapsed = int((time.time() - t0) * 1000)
        return {"response": response, "tipo": "consulta_banco", "query_executed": sql[:200] + "...", "query_results": len(data), "time_ms": elapsed, "_detail_data": data}

    # ---- VENDAS ----
    async def _handle_vendas(self, question, user_context, t0, params=None, ctx=None):
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        print(f"[SMART] Vendas params: {params}")
        periodo = params.get("periodo", "mes")
        periodo_nome = PERIODO_NOMES.get(periodo, "este mes")
        pf = _build_periodo_filter(params)
        rbac = ""
        if user_context:
            role = user_context.get("role", "vendedor")
            if role in ("admin", "diretor", "ti"):
                pass
            elif role == "gerente":
                team = user_context.get("team_codvends", [])
                if team: rbac = f" AND C.CODVEND IN ({','.join(str(c) for c in team)})"
            elif role == "vendedor":
                codvend = user_context.get("codvend", 0)
                if codvend: rbac = f" AND C.CODVEND = {codvend}"
        emp_f = f" AND UPPER(EMP.NOMEFANTASIA) LIKE UPPER('%{params['empresa']}%')" if params.get("empresa") else ""

        sql_kpis = f"SELECT COUNT(*) AS QTD_VENDAS, NVL(SUM(C.VLRNOTA),0) AS FATURAMENTO, NVL(ROUND(AVG(C.VLRNOTA),2),0) AS TICKET_MEDIO FROM TGFCAB C LEFT JOIN TSIEMP EMP ON EMP.CODEMP=C.CODEMP WHERE C.TIPMOV='V' AND C.CODTIPOPER IN (1100,1101) AND C.STATUSNOTA<>'C' {pf} {rbac} {emp_f}"
        sql_top = f"SELECT NVL(V.APELIDO,'SEM VENDEDOR') AS VENDEDOR, COUNT(*) AS QTD, NVL(SUM(C.VLRNOTA),0) AS FATURAMENTO FROM TGFCAB C LEFT JOIN TGFVEN V ON V.CODVEND=C.CODVEND LEFT JOIN TSIEMP EMP ON EMP.CODEMP=C.CODEMP WHERE C.TIPMOV='V' AND C.CODTIPOPER IN (1100,1101) AND C.STATUSNOTA<>'C' {pf} {rbac} {emp_f} GROUP BY V.APELIDO ORDER BY SUM(C.VLRNOTA) DESC"

        kr = await self.executor.execute(sql_kpis)
        if not kr.get("success"):
            return {"response": f"Erro ao consultar vendas: {kr.get('error','?')}", "tipo": "consulta_banco", "query_executed": sql_kpis[:200], "query_results": 0}
        kd = kr.get("data", [])
        if kd and isinstance(kd[0], (list, tuple)):
            cols = kr.get("columns") or ["QTD_VENDAS","FATURAMENTO","TICKET_MEDIO"]
            kd = [dict(zip(cols, row)) for row in kd]
        kpi_row = kd[0] if kd else {}

        tr = await self.executor.execute(sql_top)
        td = []
        if tr.get("success"):
            td = tr.get("data", [])
            if td and isinstance(td[0], (list, tuple)):
                tc = tr.get("columns") or ["VENDEDOR","QTD","FATURAMENTO"]
                td = [dict(zip(tc, row)) for row in td]

        fallback_response = format_vendas_response(kpi_row, periodo_nome)
        if td:
            fallback_response += "\n**Top vendedores:**\n| Vendedor | Notas | Faturamento |\n|----------|-------|-------------|\n"
            for row in td[:5]:
                if isinstance(row, dict):
                    fallback_response += f"| {str(row.get('VENDEDOR','?'))[:20]} | {fmt_num(row.get('QTD',0))} | {fmt_brl(row.get('FATURAMENTO',0))} |\n"

        if USE_LLM_NARRATOR and int(kpi_row.get("QTD_VENDAS", 0) or 0) > 0:
            summary = build_vendas_summary(kpi_row, td, periodo_nome)
            narration = await llm_narrate(question, summary, "")
            if narration:
                parts = [f"\U0001f4ca **Vendas - {periodo_nome.title()}**\n"]
                parts.append(narration)
                if td:
                    parts.append(f"\n**Top vendedores:**\n| Vendedor | Notas | Faturamento |\n|----------|-------|-------------|")
                    for row in td[:5]:
                        if isinstance(row, dict):
                            parts.append(f"| {str(row.get('VENDEDOR','?'))[:20]} | {fmt_num(row.get('QTD',0))} | {fmt_brl(row.get('FATURAMENTO',0))} |")
                response = "\n".join(parts)
            else:
                response = fallback_response
        else:
            response = fallback_response

        result_data = {"detail_data": td, "columns": ["VENDEDOR","QTD","FATURAMENTO"], "description": f"vendas - {periodo_nome}", "params": params, "intent": "vendas"}
        if ctx:
            ctx.update("vendas", params, result_data, question)
        elapsed = int((time.time() - t0) * 1000)
        return {"response": response, "tipo": "consulta_banco", "query_executed": sql_kpis[:200] + "...", "query_results": int(kpi_row.get("QTD_VENDAS",0) or 0), "time_ms": elapsed, "_detail_data": td}

    # ---- BUSCA POR CODIGO FABRICANTE ----
    async def _handle_busca_fabricante(self, question, user_context, t0, params=None, ctx=None):
        """Busca produto pelo codigo do fabricante (referencia, numfabricante, etc.)."""
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)
        cod_fab = params.get("codigo_fabricante")
        if not cod_fab:
            return {"response": "Informe o codigo do fabricante para buscar. Ex: *\"HU711/51\"* ou *\"referencia WK 950/21\"*", "tipo": "info", "query_executed": None, "query_results": None}

        print(f"[SMART] Busca fabricante: '{cod_fab}'")

        # Busca em TGFPRO (campos de referencia)
        resolved = await resolve_manufacturer_code(cod_fab, self.executor)

        # Se nao encontrou em TGFPRO, tenta em AD_TGFPROAUXMMA (auxiliares)
        if not resolved.get("found"):
            aux_result = await buscar_similares_por_codigo(cod_fab, self.executor)
            if aux_result.get("found"):
                response = format_similares(aux_result)
                elapsed = int((time.time() - t0) * 1000)
                return {"response": response, "tipo": "consulta_banco", "query_executed": f"busca auxiliar: {cod_fab}", "query_results": 1, "time_ms": elapsed}

        response = format_busca_fabricante(resolved)

        # Salvar no contexto se encontrou 1 produto
        products = resolved.get("products", [])
        if len(products) == 1 and ctx:
            result_data = {"detail_data": products, "columns": ["codprod", "produto", "marca", "referencia"], "description": f"produto {cod_fab}", "params": params, "intent": "busca_fabricante"}
            ctx.update("busca_fabricante", params, result_data, question)

        elapsed = int((time.time() - t0) * 1000)
        return {"response": response, "tipo": "consulta_banco", "query_executed": f"busca fabricante: {cod_fab}", "query_results": len(products), "time_ms": elapsed}

    # ---- SIMILARES / CROSS-REFERENCE ----
    async def _handle_similares(self, question, user_context, t0, params=None, ctx=None):
        """Busca codigos auxiliares/similares de um produto."""
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)

        codprod = params.get("codprod")
        cod_fab = params.get("codigo_fabricante")

        if not codprod and not cod_fab:
            return {"response": "Para buscar similares, informe o codigo do produto ou referencia.\nEx: *\"similares do produto 133346\"* ou *\"similares do HU711/51\"*", "tipo": "info", "query_executed": None, "query_results": None}

        # Se tem codigo fabricante mas nao codprod, resolver primeiro
        if not codprod and cod_fab:
            resolved = await resolve_manufacturer_code(cod_fab, self.executor)
            if resolved.get("found"):
                products = resolved.get("products", [])
                if len(products) == 1:
                    codprod = products[0]["codprod"]
                else:
                    # Multiplos produtos: buscar similares pelo texto auxiliar
                    sim = await buscar_similares_por_codigo(cod_fab, self.executor)
                    response = format_similares(sim)
                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": response, "tipo": "consulta_banco", "query_executed": f"similares por codigo: {cod_fab}", "query_results": len(products), "time_ms": elapsed}
            else:
                # Tenta busca por codigo auxiliar direto
                sim = await buscar_similares_por_codigo(cod_fab, self.executor)
                response = format_similares(sim)
                elapsed = int((time.time() - t0) * 1000)
                return {"response": response, "tipo": "consulta_banco", "query_executed": f"similares por codigo: {cod_fab}", "query_results": 1 if sim.get("found") else 0, "time_ms": elapsed}

        print(f"[SMART] Similares: codprod={codprod}")
        sim_data = await buscar_similares(codprod, self.executor)
        response = format_similares(sim_data)

        if ctx and sim_data.get("found"):
            result_data = {"detail_data": sim_data.get("auxiliares", []), "columns": ["codigo", "marca", "observacao"], "description": f"similares produto {codprod}", "params": params, "intent": "similares"}
            ctx.update("similares", params, result_data, question)

        elapsed = int((time.time() - t0) * 1000)
        return {"response": response, "tipo": "consulta_banco", "query_executed": f"similares: codprod={codprod}", "query_results": len(sim_data.get("auxiliares", [])), "time_ms": elapsed}

    # ---- BUSCA POR APLICACAO/VEICULO ----
    async def _handle_busca_aplicacao(self, question, user_context, t0, params=None, ctx=None):
        """Busca produtos por aplicacao/veiculo usando CARACTERISTICAS."""
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)

        aplicacao = params.get("aplicacao", "")
        if not aplicacao:
            return {"response": "Para buscar por aplicacao, informe o veiculo ou motor.\nEx: *\"pecas para scania r450\"* ou *\"filtros para motor dc13\"*", "tipo": "info", "query_executed": None, "query_results": None}

        safe_app = _safe_sql(aplicacao)
        marca_filter = ""
        if params.get("marca"):
            marca_filter = f" AND UPPER(MAR.DESCRICAO) LIKE UPPER('%{_safe_sql(params['marca'])}%')"

        sql = f"""SELECT DISTINCT PRO.CODPROD, PRO.DESCRPROD AS PRODUTO,
            NVL(MAR.DESCRICAO,'') AS MARCA,
            NVL(PRO.CARACTERISTICAS,'') AS APLICACAO, PRO.REFERENCIA
        FROM TGFPRO PRO
        LEFT JOIN TGFMAR MAR ON MAR.CODIGO = PRO.CODMARCA
        WHERE PRO.ATIVO = 'S'
          AND UPPER(PRO.CARACTERISTICAS) LIKE UPPER('%{safe_app}%')
          {marca_filter}
          AND ROWNUM <= 20
        ORDER BY MAR.DESCRICAO, PRO.DESCRPROD"""

        print(f"[SMART] Busca aplicacao: '{aplicacao}' marca={params.get('marca')}")
        result = await self.executor.execute(sql)

        if not result.get("success"):
            return {"response": f"Erro ao buscar por aplicacao: {result.get('error','?')}", "tipo": "consulta_banco", "query_executed": sql[:200], "query_results": 0}

        data = result.get("data", [])
        if data and isinstance(data[0], (list, tuple)):
            cols = result.get("columns") or ["CODPROD", "PRODUTO", "MARCA", "APLICACAO", "REFERENCIA"]
            data = [dict(zip(cols, row)) for row in data]

        if not data:
            elapsed = int((time.time() - t0) * 1000)
            return {"response": f"Nenhum produto encontrado com aplicacao **{aplicacao}**.\n\nVerifique se o nome do veiculo/motor esta correto.", "tipo": "consulta_banco", "query_executed": sql[:200], "query_results": 0, "time_ms": elapsed}

        lines = [f"\U0001f50d Encontrei **{len(data)} produto(s)** para aplicacao **{aplicacao}**"]
        if params.get("marca"):
            lines[0] += f" (marca {params['marca']})"
        lines[0] += ":\n"
        lines.append("| CodProd | Produto | Marca | Aplicacao | Ref. |")
        lines.append("|---------|---------|-------|-----------|------|")
        for r in data:
            aplic = _trunc(r.get('APLICACAO',''), 35)
            lines.append(f"| {r.get('CODPROD','')} | {str(r.get('PRODUTO',''))[:30]} | {str(r.get('MARCA',''))[:15]} | {aplic} | {str(r.get('REFERENCIA','') or '')[:15]} |")
        if len(data) >= 20:
            lines.append("\n*Mostrando os 20 primeiros resultados. Refine a busca com a marca ou nome da peca.*")
        lines.append(f"\nPara detalhes: *\"tudo sobre o produto {data[0].get('CODPROD','')}\"*")

        response = "\n".join(lines)
        elapsed = int((time.time() - t0) * 1000)

        result_data = {"detail_data": data, "columns": ["CODPROD","PRODUTO","MARCA","APLICACAO","REFERENCIA"], "description": "aplicacao", "params": params, "intent": "produto"}
        if ctx:
            ctx.update("produto", params, result_data, question)

        return {"response": response, "tipo": "consulta_banco", "query_executed": sql[:200] + "...", "query_results": len(data), "time_ms": elapsed, "_detail_data": data}

    # ---- PRODUTO 360 (visao completa) ----
    async def _handle_produto_360(self, question, user_context, t0, params=None, ctx=None):
        """Combina estoque + pendencia + vendas para visao completa de um produto."""
        if params is None:
            params = extract_entities(question, self._known_marcas, self._known_empresas, self._known_compradores)

        codprod = params.get("codprod")
        cod_fab = params.get("codigo_fabricante")
        produto_nome = params.get("produto_nome")

        # Resolver CODPROD se necessario
        if not codprod and cod_fab:
            resolved = await resolve_manufacturer_code(cod_fab, self.executor)
            if resolved.get("found"):
                products = resolved.get("products", [])
                if len(products) == 1:
                    codprod = products[0]["codprod"]
                else:
                    response = format_busca_fabricante(resolved)
                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": response, "tipo": "consulta_banco", "query_executed": f"busca: {cod_fab}", "query_results": len(products), "time_ms": elapsed}
            else:
                elapsed = int((time.time() - t0) * 1000)
                return {"response": f"Nenhum produto encontrado com o codigo **{cod_fab}**.", "tipo": "consulta_banco", "query_executed": f"busca: {cod_fab}", "query_results": 0, "time_ms": elapsed}

        if not codprod and produto_nome:
            sql_find = f"SELECT PRO.CODPROD, PRO.DESCRPROD, NVL(PRO.CARACTERISTICAS,'') AS APLICACAO FROM TGFPRO PRO WHERE UPPER(PRO.DESCRPROD) LIKE UPPER('%{produto_nome.replace(chr(39), chr(39)+chr(39))}%') AND PRO.ATIVO='S' AND ROWNUM<=5"
            r_find = await self.executor.execute(sql_find)
            if r_find.get("success") and r_find.get("data"):
                fdata = r_find["data"]
                if fdata and isinstance(fdata[0], (list, tuple)):
                    fdata = [dict(zip(["CODPROD", "DESCRPROD", "APLICACAO"], row)) for row in fdata]
                if len(fdata) == 1:
                    codprod = int(fdata[0].get("CODPROD", 0) or 0)
                else:
                    lines = [f"Encontrei **{len(fdata)} produtos** com '{produto_nome}':\n"]
                    lines.append("| CodProd | Produto | Aplicacao |")
                    lines.append("|---------|---------|-----------|")
                    for r in fdata:
                        aplic = _trunc(r.get('APLICACAO',''), 40)
                        lines.append(f"| {r.get('CODPROD','')} | {str(r.get('DESCRPROD',''))[:50]} | {aplic} |")
                    lines.append(f"\nEspecifique: *\"tudo sobre o produto {fdata[0].get('CODPROD','')}\"*")
                    elapsed = int((time.time() - t0) * 1000)
                    return {"response": "\n".join(lines), "tipo": "consulta_banco", "query_executed": sql_find[:200], "query_results": len(fdata), "time_ms": elapsed}

        if not codprod:
            return {"response": "Para ver a visao 360, informe o codigo do produto.\nEx: *\"tudo sobre o produto 133346\"*", "tipo": "info", "query_executed": None, "query_results": None}

        print(f"[SMART] Produto 360: codprod={codprod}")

        # Executar consultas em paralelo
        sql_prod = f"""SELECT PRO.CODPROD, PRO.DESCRPROD AS PRODUTO, NVL(MAR.DESCRICAO,'') AS MARCA,
            PRO.REFERENCIA, PRO.AD_NUMFABRICANTE, PRO.NCM, PRO.CODVOL,
            NVL(PRO.CARACTERISTICAS,'') AS APLICACAO,
            NVL(PRO.COMPLDESC,'') AS COMPLEMENTO,
            NVL(PRO.AD_NUMORIGINAL,'') AS NUM_ORIGINAL,
            NVL(PRO.REFFORN,'') AS REF_FORNECEDOR
        FROM TGFPRO PRO LEFT JOIN TGFMAR MAR ON MAR.CODIGO = PRO.CODMARCA
        WHERE PRO.CODPROD = {codprod}"""

        sql_est = f"""SELECT EMP.NOMEFANTASIA AS EMPRESA, NVL(E.ESTOQUE,0) AS ESTOQUE, NVL(E.ESTMIN,0) AS ESTMIN
        FROM TGFEST E
        LEFT JOIN TSIEMP EMP ON EMP.CODEMP = E.CODEMP
        WHERE E.CODPROD = {codprod} AND E.CODLOCAL = 0 AND NVL(E.ATIVO,'S') = 'S'
        ORDER BY EMP.NOMEFANTASIA"""

        sql_vendas = f"""SELECT COUNT(DISTINCT C.NUNOTA) AS QTD_VENDAS, SUM(I.QTDNEG) AS QTD_VENDIDA,
            SUM(I.VLRTOT) AS VLR_TOTAL
        FROM TGFITE I JOIN TGFCAB C ON C.NUNOTA = I.NUNOTA
        WHERE I.CODPROD = {codprod}
          AND C.TIPMOV = 'V' AND C.CODTIPOPER IN (1100,1101) AND C.STATUSNOTA <> 'C'
          AND C.DTNEG >= ADD_MONTHS(TRUNC(SYSDATE), -3)"""

        # Executar em paralelo
        r_prod, r_est, r_vendas = await asyncio.gather(
            self.executor.execute(sql_prod),
            self.executor.execute(sql_est),
            self.executor.execute(sql_vendas),
        )

        # Montar info do produto
        prod_info = {"codprod": codprod, "produto": "?", "marca": "", "referencia": ""}
        if r_prod.get("success") and r_prod.get("data"):
            pdata = r_prod["data"]
            if pdata and isinstance(pdata[0], (list, tuple)):
                cols = r_prod.get("columns") or ["CODPROD", "PRODUTO", "MARCA", "REFERENCIA", "AD_NUMFABRICANTE", "NCM", "CODVOL", "APLICACAO", "COMPLEMENTO", "NUM_ORIGINAL", "REF_FORNECEDOR"]
                pdata = [dict(zip(cols, row)) for row in pdata]
            if pdata and isinstance(pdata[0], dict):
                p = pdata[0]
                prod_info["produto"] = str(p.get("PRODUTO", "?"))
                prod_info["marca"] = str(p.get("MARCA", ""))
                prod_info["referencia"] = str(p.get("REFERENCIA", "") or p.get("AD_NUMFABRICANTE", "") or "")
                prod_info["aplicacao"] = str(p.get("APLICACAO", "") or "")
                prod_info["complemento"] = str(p.get("COMPLEMENTO", "") or "")
                prod_info["num_original"] = str(p.get("NUM_ORIGINAL", "") or "")
                prod_info["ref_fornecedor"] = str(p.get("REF_FORNECEDOR", "") or "")

        # Estoque
        estoque_data = []
        if r_est.get("success") and r_est.get("data"):
            edata = r_est["data"]
            if edata and isinstance(edata[0], (list, tuple)):
                cols = r_est.get("columns") or ["EMPRESA", "ESTOQUE", "ESTMIN"]
                edata = [dict(zip(cols, row)) for row in edata]
            estoque_data = [r for r in edata if isinstance(r, dict)]

        # Pendencia de compras (usa handler existente internamente)
        pend_params = {"codprod": codprod}
        sql_pend = f"""SELECT CAB.NUNOTA AS PEDIDO,
            CASE WHEN CAB.CODTIPOPER = 1313 THEN 'Casada' WHEN CAB.CODTIPOPER = 1301 THEN 'Estoque' END AS TIPO_COMPRA,
            PAR.NOMEPARC AS FORNECEDOR, SUM(ITE.QTDNEG - NVL(ITE.QTDENTREGUE,0)) AS QTD_PENDENTE,
            SUM((ITE.QTDNEG - NVL(ITE.QTDENTREGUE,0)) * ITE.VLRUNIT) AS VLR_PENDENTE,
            CASE WHEN CAB.AD_DTPREVCHEG IS NULL THEN 'SEM PREVISAO'
                 WHEN CAB.AD_DTPREVCHEG < TRUNC(SYSDATE) THEN 'ATRASADO'
                 ELSE 'NO PRAZO' END AS STATUS_ENTREGA
        FROM TGFCAB CAB
        JOIN TGFITE ITE ON ITE.NUNOTA = CAB.NUNOTA
        JOIN TGFPRO PRO ON PRO.CODPROD = ITE.CODPROD
        JOIN TGFPAR PAR ON PAR.CODPARC = CAB.CODPARC
        WHERE CAB.CODTIPOPER IN (1301, 1313) AND CAB.PENDENTE = 'S'
          AND ITE.CODPROD = {codprod}
          AND (ITE.QTDNEG - NVL(ITE.QTDENTREGUE,0)) > 0
        GROUP BY CAB.NUNOTA, CAB.CODTIPOPER, PAR.NOMEPARC, CAB.AD_DTPREVCHEG
        ORDER BY CAB.NUNOTA"""

        r_pend = await self.executor.execute(sql_pend)
        pendencia_data = {"detail_data": []}
        if r_pend.get("success") and r_pend.get("data"):
            pd_data = r_pend["data"]
            if pd_data and isinstance(pd_data[0], (list, tuple)):
                cols = r_pend.get("columns") or ["PEDIDO", "TIPO_COMPRA", "FORNECEDOR", "QTD_PENDENTE", "VLR_PENDENTE", "STATUS_ENTREGA"]
                pd_data = [dict(zip(cols, row)) for row in pd_data]
            pendencia_data["detail_data"] = [r for r in pd_data if isinstance(r, dict)]

        # Vendas
        vendas_info = {}
        if r_vendas.get("success") and r_vendas.get("data"):
            vdata = r_vendas["data"]
            if vdata and isinstance(vdata[0], (list, tuple)):
                cols = r_vendas.get("columns") or ["QTD_VENDAS", "QTD_VENDIDA", "VLR_TOTAL"]
                vdata = [dict(zip(cols, row)) for row in vdata]
            if vdata and isinstance(vdata[0], dict):
                vendas_info = vdata[0]

        # Formatar resposta
        response = format_produto_360(prod_info, estoque_data, pendencia_data, vendas_info)

        # Salvar no contexto
        if ctx:
            all_data = estoque_data + pendencia_data.get("detail_data", [])
            result_data = {"detail_data": all_data, "columns": ["CODPROD"], "description": f"produto 360 - {codprod}", "params": params, "intent": "produto_360"}
            ctx.update("produto_360", params, result_data, question)

        elapsed = int((time.time() - t0) * 1000)
        return {"response": response, "tipo": "consulta_banco", "query_executed": f"produto 360: codprod={codprod}", "query_results": 1, "time_ms": elapsed}

    # ---- EXCEL ----
    async def _handle_excel_followup(self, user_context, ctx=None):
        last_result = ctx.last_result if ctx else {}
        if not last_result or not last_result.get("detail_data"):
            return {"response": "Nao tenho dados para gerar o arquivo. Faca uma consulta primeiro.", "tipo": "info", "query_executed": None, "query_results": None}
        data = last_result["detail_data"]
        columns = last_result["columns"]
        description = last_result["description"]
        params = last_result.get("params", {})
        intent = last_result.get("intent", "dados")
        fn = params.get("marca") or params.get("fornecedor") or params.get("empresa") or params.get("periodo", "geral")
        fn = re.sub(r'[^\w\s-]', '', str(fn)).strip().replace(' ', '_')
        filename = f"{intent}_{fn}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        try:
            fp = generate_excel(data=data, columns=columns, filename=filename, title=description.title())
            rp = Path(fp).name
            url = f"/static/exports/{rp}"
            s = "s" if len(data) > 1 else ""
            r = f"\U0001f4ca **Arquivo gerado com sucesso!**\n\n**{len(data)} registro{s}** exportado{s}.\n\n\U0001f4e5 **[Clique aqui para baixar]({url})**\n\n*Arquivo: {rp}*"
            return {"response": r, "tipo": "arquivo", "query_executed": None, "query_results": len(data), "download_url": url}
        except Exception as e:
            return {"response": f"Erro ao gerar arquivo: {str(e)}", "tipo": "erro", "query_executed": None, "query_results": None}
