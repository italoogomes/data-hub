"""
MMarra Data Hub - Gerador de Excel/CSV.
Exporta dados para formatos de planilha.
Extraído de smart_agent.py na refatoração modular.
"""

import os
from pathlib import Path
from datetime import datetime


def generate_excel(data, columns, filename, title=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return generate_csv(data, columns, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    hfill = PatternFill(start_color="0E75B9", end_color="0E75B9", fill_type="solid")
    hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    dfont = Font(name="Arial", size=9)
    brd = Border(bottom=Side(style="thin", color="E0E0E0"))
    sr = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.cell(row=1, column=1, value=title).font = Font(name="Arial", size=12, bold=True, color="0E75B9")
        ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=8, color="888888")
        sr = 4
    for ci, cn in enumerate(columns, 1):
        c = ws.cell(row=sr, column=ci, value=cn); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
    ccols = {"VLR_UNITARIO","VLR_PENDENTE","VLR_TOTAL_PENDENTE","VALOR","FATURAMENTO","TICKET_MEDIO"}
    ncols = {"QTD_PEDIDA","QTD_ATENDIDA","QTD_PENDENTE","DIAS_ABERTO","CODPROD","PEDIDO","ESTOQUE","ESTMIN","QTD"}
    for ri, rd in enumerate(data, sr + 1):
        for ci, cn in enumerate(columns, 1):
            val = rd.get(cn, "") if isinstance(rd, dict) else (rd[ci-1] if ci <= len(rd) else "")
            cell = ws.cell(row=ri, column=ci, value=val); cell.font = dfont; cell.border = brd
            if cn in ccols:
                try: cell.value = float(val or 0); cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                except: pass
            elif cn in ncols:
                try: cell.value = int(float(val or 0)); cell.alignment = Alignment(horizontal="right")
                except: pass
    for ci, cn in enumerate(columns, 1):
        ml = len(str(cn))
        for ri in range(sr+1, min(sr+50, len(data)+sr+1)):
            cv = ws.cell(row=ri, column=ci).value
            if cv: ml = max(ml, min(len(str(cv)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = ml + 3
    ws.auto_filter.ref = f"A{sr}:{get_column_letter(len(columns))}{sr + len(data)}"
    ws.freeze_panes = ws.cell(row=sr + 1, column=1)
    static_dir = Path(__file__).parent.parent / "api" / "static" / "exports"
    static_dir.mkdir(parents=True, exist_ok=True)
    fp = static_dir / filename; wb.save(str(fp)); return str(fp)


def generate_csv(data, columns, filename):
    static_dir = Path(__file__).parent.parent / "api" / "static" / "exports"
    static_dir.mkdir(parents=True, exist_ok=True)
    fp = static_dir / filename.replace(".xlsx", ".csv")
    with open(fp, "w", encoding="utf-8-sig") as f:
        f.write(";".join(columns) + "\n")
        for row in data:
            vals = [str(row.get(c,"")).replace(";",",") for c in columns] if isinstance(row, dict) else [str(v).replace(";",",") for v in row]
            f.write(";".join(vals) + "\n")
    return str(fp)
